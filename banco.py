import os
import secrets
import time
import unicodedata
import base64
import streamlit as st
import psycopg2
from psycopg2 import pool
from psycopg2.extras import Json
from datetime import datetime, date, timedelta
from zoneinfo import ZoneInfo


# ==================================================
# CONTROLE DE VERIFICAÇÃO DE CONEXÕES (ver conectar())
# ==================================================

_CONEXOES_VERIFICADAS = {}
_VALIDADE_VERIFICACAO = 30


# ==================================================
# CONFIGURAÇÃO SUPABASE
# ==================================================

HOST = os.getenv("SUPABASE_HOST")
PORT = os.getenv("SUPABASE_PORT")
DATABASE = os.getenv("SUPABASE_DATABASE")

USER = os.getenv("SUPABASE_USER")
PASSWORD = os.getenv("SUPABASE_PASSWORD")

# ==================================================
# RUAS PADRÃO (usadas só para semear um armazém novo)
# ==================================================

RUAS_PADRAO = [
    "Rua 01",
    "Rua 02",
    "Rua 03",
    "Rua 04",
    "Rua 05",
    "Rua 06",
    "Rua 07",
    "Rua 35&32",
    "Rua 33&34"
]

# ==================================================
# FUNDADOR (SECRETS)
# ==================================================

USUARIO_FUNDADOR = os.getenv(
    "FUNDADOR_USUARIO"
)

SENHA_FUNDADOR = os.getenv(
    "FUNDADOR_SENHA"
)

# ==================================================
# CONEXÃO (POOL REAPROVEITADO)
# ==================================================
# Antes, cada ação abria uma conexão nova com o Supabase
# do zero (handshake TLS completo a cada clique), o que
# deixava tudo mais lento. Agora um pool mantém conexões
# já abertas prontas para uso — conectar() pega uma
# emprestada, liberar() devolve pro pool (sem fechar de
# verdade), evitando repetir esse custo a cada ação.
#
# IMPORTANTE (limite do Supabase): o pooler do Supabase em
# "session mode" (porta 5432) aceita no máximo 15 conexões
# simultâneas no total — esse limite é do Postgres/pooler,
# não do psycopg2. Por isso o maxconn do pool abaixo fica
# em 10 (com folga), e não em 15 ou mais: se o app tiver
# mais de um processo rodando ao mesmo tempo (ex.: durante
# um redeploy, o processo antigo e o novo convivendo por
# alguns segundos), CADA processo cria o seu próprio pool
# — os limites não são compartilhados entre processos, e
# só a soma de todos eles é que precisa caber nos 15.

@st.cache_resource(show_spinner=False)
def _obter_pool():

    # ThreadedConnectionPool (não SimpleConnectionPool): o pool é
    # compartilhado por TODAS as sessões/usuários do app ao mesmo
    # tempo (é um recurso cacheado a nível de processo). O
    # SimpleConnectionPool não é seguro para uso concorrente por
    # várias threads — com vários usuários mexendo no app ao mesmo
    # tempo, isso podia causar lentidão e travamentos aleatórios.

    return psycopg2.pool.ThreadedConnectionPool(
        1,
        10,
        host=HOST,
        port=PORT,
        database=DATABASE,
        user=USER,
        password=PASSWORD,
        connect_timeout=10
    )


def conectar():

    # O Supabase (assim como bancos gerenciados em geral) derruba
    # conexões que ficam muito tempo ociosas no pool, mas o psycopg2
    # só percebe isso na hora de usar a conexão — daí o erro
    # "server closed the connection unexpectedly". Antes, o teste
    # (SELECT 1) rodava em TODA chamada a conectar() — ou seja, toda
    # ação do app pagava duas idas ao banco (uma pra testar, outra
    # pra fazer o que interessa). Agora cada conexão só é retestada
    # se já faz mais de _VALIDADE_VERIFICACAO segundos desde a
    # última vez que ELA MESMA foi verificada (é por isso que a
    # chave é id(conn): o pool reaproveita os mesmos objetos de
    # conexão, então dá pra "lembrar" quais já foram checados
    # recentemente sem precisar checar de novo a cada clique).

    pool_obj = _obter_pool()
    conn = pool_obj.getconn()

    agora = time.monotonic()
    chave_conexao = id(conn)
    verificada_em = _CONEXOES_VERIFICADAS.get(chave_conexao, 0)

    if agora - verificada_em > _VALIDADE_VERIFICACAO:

        try:
            with conn.cursor() as cursor:
                cursor.execute("SELECT 1")

            _CONEXOES_VERIFICADAS[chave_conexao] = agora

        except Exception:

            _CONEXOES_VERIFICADAS.pop(chave_conexao, None)

            try:
                pool_obj.putconn(conn, close=True)
            except Exception:
                pass

            conn = pool_obj.getconn()
            _CONEXOES_VERIFICADAS[id(conn)] = agora

    return conn


def liberar(conn):

    try:
        _obter_pool().putconn(conn)
    except Exception:
        pass


@st.cache_resource(show_spinner=False)
def _garantir_schema():

    print("🔎 Iniciando conexão com o banco...", flush=True)
    print(f"HOST={HOST!r} PORT={PORT!r} DATABASE={DATABASE!r} USER={USER!r}", flush=True)

    try:
        conn = conectar()
        print("✅ Conectado com sucesso.", flush=True)
    except Exception as e:
        print(f"❌ ERRO AO CONECTAR: {e}", flush=True)
        raise

    try:
        cursor = conn.cursor()

        # ==============================
        # NOTAS
        # ==============================

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS notas (
            rua TEXT PRIMARY KEY,
            nota REAL,
            dupla TEXT
        )
        """)

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS historico_notas (
            id BIGSERIAL PRIMARY KEY,
            rua TEXT NOT NULL,
            nota REAL,
            dupla TEXT,
            data_atualizacao TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """)

        # ==============================
        # REMANEJAMENTO
        # ==============================

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS remanejamento (
            id BIGSERIAL PRIMARY KEY,
            item TEXT NOT NULL,
            prioridade TEXT DEFAULT 'Normal'
        )
        """)

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS historico_remanejamento (
            id BIGSERIAL PRIMARY KEY,
            item TEXT NOT NULL,
            prioridade TEXT,
            data_hora TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """)

        # ==============================
        # REMANEJAMENTO AGENDADO (por horário/dia da semana)
        # ==============================
        # Itens que entram e saem do painel sozinhos, de acordo com o
        # horário e os dias da semana configurados — sem precisar ser
        # criados/apagados na mão. dias_semana usa a convenção do Python
        # (date.weekday()): 0=Segunda ... 6=Domingo.

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS remanejamento_agendado (
            id BIGSERIAL PRIMARY KEY,
            armazem_id BIGINT NOT NULL REFERENCES armazens(id),
            item TEXT NOT NULL,
            prioridade TEXT DEFAULT 'Normal',
            hora_inicio TIME NOT NULL,
            hora_fim TIME NOT NULL,
            dias_semana INTEGER[] NOT NULL,
            criado_por TEXT,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """)

        cursor.execute("CREATE INDEX IF NOT EXISTS idx_reman_agendado_armazem ON remanejamento_agendado (armazem_id)")

        # ==============================
        # SAC
        # ==============================

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS sac_historico (
            mes_ano TEXT PRIMARY KEY,
            reclamacoes INTEGER,
            meta INTEGER
        )
        """)

        # ==============================
        # ANÁLISE TÉCNICA (SAC)
        # ==============================

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS analise_tecnica (
            id BIGSERIAL PRIMARY KEY,
            nome TEXT,
            tipo_erro TEXT NOT NULL,
            data_erro DATE NOT NULL,
            data_fechamento DATE,
            descricao TEXT,
            chamado TEXT,
            cliente TEXT,
            nota_fiscal TEXT,
            cod_produto TEXT,
            produto TEXT,
            tratativa TEXT,
            hora TIME,
            separador TEXT,
            volume TEXT,
            carga TEXT,
            regiao TEXT,
            motorista TEXT,
            balanca TEXT,
            conferente TEXT,
            vinculos_notificados JSONB DEFAULT '[]'::jsonb,
            registrado_por TEXT,
            data_registro TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """)

        # ==============================
        # AUDITORIA DE ATIVIDADES
        # ==============================

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS auditoria_atividades (
            id BIGSERIAL PRIMARY KEY,
            nome TEXT NOT NULL,
            funcao TEXT NOT NULL,
            qtd_acertos INTEGER DEFAULT 0,
            qtd_erros INTEGER DEFAULT 0,
            data_atividade DATE NOT NULL,
            descricao TEXT,
            registrado_por TEXT,
            data_registro TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """)

        # ==============================
        # RODÍZIO - FIM DE EXPEDIENTE
        # ==============================

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS rotativo_pessoas (
            id BIGSERIAL PRIMARY KEY,
            nome TEXT UNIQUE NOT NULL,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """)

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS rotativo_atividades (
            id BIGSERIAL PRIMARY KEY,
            nome TEXT UNIQUE NOT NULL,
            tipo TEXT DEFAULT 'rotativo',
            pessoa_fixa TEXT,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """)

        # ==============================
        # CHECKLISTS
        # ==============================

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS checklist_hidraulicos (
            id BIGSERIAL PRIMARY KEY,
            nome TEXT NOT NULL,
            numero TEXT NOT NULL,
            data_checklist DATE NOT NULL,
            status TEXT NOT NULL,
            descricao TEXT,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """)

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS checklist_carrinhos (
            id BIGSERIAL PRIMARY KEY,
            nome TEXT NOT NULL,
            numero TEXT NOT NULL,
            data_checklist DATE NOT NULL,
            status TEXT NOT NULL,
            descricao TEXT,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """)

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS checklist_empilhadeiras (
            id BIGSERIAL PRIMARY KEY,
            nome TEXT NOT NULL,
            numero TEXT NOT NULL,
            data_checklist DATE NOT NULL,
            status TEXT NOT NULL,
            descricao TEXT,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """)

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS checklist_pigmentacao (
            id BIGSERIAL PRIMARY KEY,
            nome TEXT NOT NULL,
            data_checklist DATE NOT NULL,
            status TEXT NOT NULL,
            descricao TEXT,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """)

        # ==============================
        # EQUIPAMENTOS: RESPONSÁVEIS E CARRINHOS FIXOS
        # ==============================

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS responsaveis_hidraulicos (
            id BIGSERIAL PRIMARY KEY,
            nome TEXT NOT NULL,
            numero TEXT NOT NULL,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """)

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS responsaveis_carrinhos (
            id BIGSERIAL PRIMARY KEY,
            nome TEXT NOT NULL,
            numero TEXT NOT NULL,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """)

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS carrinhos_fixos (
            id BIGSERIAL PRIMARY KEY,
            local TEXT NOT NULL,
            numero TEXT NOT NULL,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """)

        # Carrinhos fixos por local já vêm pré-cadastrados na primeira
        # vez que o app roda — depois disso, tudo é editável normalmente
        # pela aba de Equipamentos no Administrativo.

        cursor.execute("SELECT COUNT(*) FROM carrinhos_fixos")

        if cursor.fetchone()[0] == 0:

            carrinhos_padrao = (
                [("Remanejamento", numero) for numero in ["06", "08", "09", "13"]] +
                [("Fracionado", numero) for numero in ["10", "11", "12"]]
            )

            for local, numero in carrinhos_padrao:

                cursor.execute("""
                INSERT INTO carrinhos_fixos (local, numero)
                VALUES (%s, %s)
                """, (local, numero))

        # ==============================
        # USUÁRIOS
        # ==============================

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS usuarios (
            id BIGSERIAL PRIMARY KEY,
            usuario TEXT UNIQUE,
            senha TEXT,
            tipo TEXT DEFAULT 'usuario',
            trocar_senha INTEGER DEFAULT 1
        )
        """)

        # ==============================
        # SESSÕES ATIVAS (login persistente)
        # ==============================
        # Guarda um token por login, associado a um link (URL) que o
        # navegador mantém mesmo depois de um F5 ou de uma queda de
        # rede rápida — assim, ao reconectar, o app reconhece o token
        # e restaura a sessão sem pedir login de novo.

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS sessoes_ativas (
            token TEXT PRIMARY KEY,
            usuario TEXT NOT NULL,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            expira_em TIMESTAMP
        )
        """)

        conn.commit()

    finally:
        liberar(conn)

    # ==============================
    # MIGRAÇÃO: colunas de auditoria
    # (seguro rodar mesmo com o banco já existente)
    # ==============================

    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("ALTER TABLE notas ADD COLUMN IF NOT EXISTS atualizado_por TEXT")
        cursor.execute("ALTER TABLE historico_notas ADD COLUMN IF NOT EXISTS usuario TEXT")
        cursor.execute("ALTER TABLE remanejamento ADD COLUMN IF NOT EXISTS criado_por TEXT")
        cursor.execute("ALTER TABLE historico_remanejamento ADD COLUMN IF NOT EXISTS usuario TEXT")
        cursor.execute("ALTER TABLE sac_historico ADD COLUMN IF NOT EXISTS atualizado_por TEXT")
        cursor.execute("ALTER TABLE sac_historico ADD COLUMN IF NOT EXISTS atualizado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP")
        cursor.execute("ALTER TABLE analise_tecnica ALTER COLUMN nome DROP NOT NULL")
        cursor.execute("ALTER TABLE analise_tecnica ADD COLUMN IF NOT EXISTS descricao TEXT")
        cursor.execute("ALTER TABLE analise_tecnica ADD COLUMN IF NOT EXISTS chamado TEXT")
        cursor.execute("ALTER TABLE analise_tecnica ADD COLUMN IF NOT EXISTS cliente TEXT")
        cursor.execute("ALTER TABLE analise_tecnica ADD COLUMN IF NOT EXISTS nota_fiscal TEXT")
        cursor.execute("ALTER TABLE analise_tecnica ADD COLUMN IF NOT EXISTS cod_produto TEXT")
        cursor.execute("ALTER TABLE analise_tecnica ADD COLUMN IF NOT EXISTS produto TEXT")
        cursor.execute("ALTER TABLE analise_tecnica ADD COLUMN IF NOT EXISTS tratativa TEXT")
        cursor.execute("ALTER TABLE analise_tecnica ADD COLUMN IF NOT EXISTS hora TIME")
        cursor.execute("ALTER TABLE analise_tecnica ADD COLUMN IF NOT EXISTS separador TEXT")
        cursor.execute("ALTER TABLE analise_tecnica ADD COLUMN IF NOT EXISTS volume TEXT")
        cursor.execute("ALTER TABLE analise_tecnica ADD COLUMN IF NOT EXISTS carga TEXT")
        cursor.execute("ALTER TABLE analise_tecnica ADD COLUMN IF NOT EXISTS regiao TEXT")
        cursor.execute("ALTER TABLE analise_tecnica ADD COLUMN IF NOT EXISTS motorista TEXT")
        cursor.execute("ALTER TABLE analise_tecnica ADD COLUMN IF NOT EXISTS balanca TEXT")
        cursor.execute("ALTER TABLE analise_tecnica ADD COLUMN IF NOT EXISTS conferente TEXT")
        cursor.execute("ALTER TABLE analise_tecnica ADD COLUMN IF NOT EXISTS vinculos_notificados JSONB DEFAULT '[]'::jsonb")
        cursor.execute("ALTER TABLE analise_tecnica ADD COLUMN IF NOT EXISTS data_fechamento DATE")
        cursor.execute("ALTER TABLE rotativo_atividades ADD COLUMN IF NOT EXISTS tipo TEXT DEFAULT 'rotativo'")
        cursor.execute("ALTER TABLE rotativo_atividades ADD COLUMN IF NOT EXISTS pessoa_fixa TEXT")

        # Rastreio de "quem está logado agora": guarda o momento do último
        # acesso/atividade de cada usuário, atualizado no login e a cada
        # renovação automática do rodapé (ver render_status_footer no app.py).
        cursor.execute("ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS ultimo_acesso TIMESTAMP")

        # Manutenção de equipamentos (checklist de hidráulicos, carrinhos e
        # empilhadeiras): permite marcar um item "Não Conforme" como enviado
        # para manutenção, e depois registrar o retorno.
        for tabela_checklist in ("checklist_hidraulicos", "checklist_carrinhos", "checklist_empilhadeiras"):

            cursor.execute(f"ALTER TABLE {tabela_checklist} ADD COLUMN IF NOT EXISTS em_manutencao BOOLEAN DEFAULT FALSE")
            cursor.execute(f"ALTER TABLE {tabela_checklist} ADD COLUMN IF NOT EXISTS manutencao_motivo TEXT")
            cursor.execute(f"ALTER TABLE {tabela_checklist} ADD COLUMN IF NOT EXISTS manutencao_enviado_por TEXT")
            cursor.execute(f"ALTER TABLE {tabela_checklist} ADD COLUMN IF NOT EXISTS manutencao_enviado_em TIMESTAMP")
            cursor.execute(f"ALTER TABLE {tabela_checklist} ADD COLUMN IF NOT EXISTS manutencao_retornado_por TEXT")
            cursor.execute(f"ALTER TABLE {tabela_checklist} ADD COLUMN IF NOT EXISTS manutencao_retornado_em TIMESTAMP")

        conn.commit()

    finally:
        liberar(conn)

    # ==============================
    # MIGRAÇÃO: MULTI-ARMAZÉM
    # ==============================
    # Cada cliente/armazém que usar o Luxiz IA passa a ter os
    # próprios dados isolados dentro do MESMO banco, através da
    # coluna armazem_id. Os dados que já existiam (de antes dessa
    # migração) são automaticamente colocados no "Armazém Principal"
    # — nada se perde.

    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS armazens (
            id BIGSERIAL PRIMARY KEY,
            nome TEXT UNIQUE NOT NULL,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """)

        cursor.execute("SELECT id FROM armazens ORDER BY id ASC LIMIT 1")
        linha_armazem_padrao = cursor.fetchone()

        if not linha_armazem_padrao:

            cursor.execute("""
            INSERT INTO armazens (nome)
            VALUES (%s)
            RETURNING id
            """, ("Armazém Principal",))

            linha_armazem_padrao = cursor.fetchone()

        ID_ARMAZEM_PADRAO = linha_armazem_padrao[0]

        # ==============================
        # RUAS (cadastro dinâmico por armazém)
        # ==============================
        # Antes a lista de ruas era fixa (9 ruas "hardcoded" no código).
        # Agora cada armazém tem seu próprio cadastro de ruas, editável
        # no Administrativo > Dashboard ("Criar rua" / "Excluir rua").

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS ruas (
            id BIGSERIAL PRIMARY KEY,
            armazem_id BIGINT REFERENCES armazens(id),
            nome TEXT NOT NULL,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE (armazem_id, nome)
        )
        """)

        cursor.execute("CREATE INDEX IF NOT EXISTS idx_ruas_armazem ON ruas (armazem_id)")

        # Semeia as 9 ruas padrão em qualquer armazém que ainda não
        # tenha nenhuma rua cadastrada — preserva o comportamento atual
        # sem perder nada para quem já usa o app.
        cursor.execute("SELECT id FROM armazens")
        ids_todos_armazens = [linha[0] for linha in cursor.fetchall()]

        for id_armazem in ids_todos_armazens:

            cursor.execute(
                "SELECT COUNT(*) FROM ruas WHERE armazem_id = %s",
                (id_armazem,)
            )

            if cursor.fetchone()[0] == 0:

                for nome_rua in RUAS_PADRAO:

                    cursor.execute(
                        """
                        INSERT INTO ruas (armazem_id, nome)
                        VALUES (%s, %s)
                        ON CONFLICT (armazem_id, nome) DO NOTHING
                        """,
                        (id_armazem, nome_rua)
                    )

        conn.commit()

        # ==============================
        # CONTROLE DE EPI's
        # ==============================

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS epis (
            id BIGSERIAL PRIMARY KEY,
            armazem_id BIGINT REFERENCES armazens(id),
            nome TEXT NOT NULL,
            epi TEXT NOT NULL,
            data DATE NOT NULL,
            assinatura TEXT,
            assinado_por TEXT,
            assinado_em TIMESTAMP,
            criado_por TEXT,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """)

        cursor.execute("CREATE INDEX IF NOT EXISTS idx_epis_armazem ON epis (armazem_id)")

        # ==============================
        # NOTIFICAÇÕES (lidas / excluídas)
        # ==============================
        # "notificacoes_lidas" é individual: cada usuário marca a sua
        # como lida, sem afetar os outros. "notificacoes_excluidas" é
        # global: quando a Gestão/Fundador exclui, some para todo mundo.

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS notificacoes_lidas (
            usuario TEXT NOT NULL,
            notificacao_id TEXT NOT NULL,
            armazem_id BIGINT REFERENCES armazens(id),
            lida_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            PRIMARY KEY (usuario, notificacao_id, armazem_id)
        )
        """)

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS notificacoes_excluidas (
            notificacao_id TEXT NOT NULL,
            armazem_id BIGINT REFERENCES armazens(id),
            excluida_por TEXT,
            excluida_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            PRIMARY KEY (notificacao_id, armazem_id)
        )
        """)

        cursor.execute("CREATE INDEX IF NOT EXISTS idx_notif_lidas_usuario ON notificacoes_lidas (usuario, armazem_id)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_notif_excluidas_armazem ON notificacoes_excluidas (armazem_id)")

        conn.commit()

        # ==============================
        # PERFIS (nome, sobrenome, função e foto de cada usuário)
        # ==============================
        # Um perfil por usuário (usuario é PK e referencia a tabela
        # usuarios). O objetivo é resolver a ambiguidade de nomes
        # repetidos (ex.: dois "João") — em vez de cadastrar registros
        # (EPI, auditoria etc.) só com o primeiro nome, o sistema passa
        # a reconhecer nome + sobrenome de cada pessoa.

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS perfis (
            usuario TEXT PRIMARY KEY REFERENCES usuarios(usuario) ON DELETE CASCADE,
            nome TEXT NOT NULL,
            sobrenome TEXT NOT NULL,
            funcao TEXT,
            foto TEXT,
            armazem_id BIGINT REFERENCES armazens(id),
            atualizado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """)

        cursor.execute("CREATE INDEX IF NOT EXISTS idx_perfis_armazem ON perfis (armazem_id)")

        conn.commit()

        # ==============================
        # PLANILHA DO SAC (vinculada pelo Administrativo)
        # ==============================
        # PLANILHA VINCULADA DO SAC
        # ==============================
        # Guarda o último arquivo Excel enviado pelo Administrativo,
        # por armazém. O conteúdo fica salvo no próprio banco (BYTEA)
        # — o app relê ESSA cópia periodicamente, não o arquivo no
        # computador da pessoa (que ele não teria como acessar).

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS planilha_sac (
            armazem_id BIGINT PRIMARY KEY REFERENCES armazens(id),
            nome_arquivo TEXT,
            conteudo BYTEA,
            enviado_por TEXT,
            enviado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            ultima_checagem TIMESTAMP,
            ultimo_resultado TEXT
        )
        """)

        conn.commit()

        TABELAS_COM_ARMAZEM = [
            "notas", "historico_notas",
            "remanejamento", "historico_remanejamento",
            "sac_historico",
            "analise_tecnica",
            "auditoria_atividades",
            "rotativo_pessoas", "rotativo_atividades",
            "checklist_hidraulicos", "checklist_carrinhos",
            "checklist_empilhadeiras", "checklist_pigmentacao",
            "responsaveis_hidraulicos", "responsaveis_carrinhos",
            "carrinhos_fixos",
            "usuarios",
        ]

        for tabela in TABELAS_COM_ARMAZEM:

            cursor.execute(
                f"ALTER TABLE {tabela} ADD COLUMN IF NOT EXISTS "
                f"armazem_id BIGINT REFERENCES armazens(id)"
            )

            cursor.execute(
                f"UPDATE {tabela} SET armazem_id = %s WHERE armazem_id IS NULL",
                (ID_ARMAZEM_PADRAO,)
            )

            cursor.execute(
                f"ALTER TABLE {tabela} ALTER COLUMN armazem_id SET NOT NULL"
            )

        # notas: cada armazém tem sua própria "Rua 01", "Rua 02"... então
        # a chave única passa a ser (armazem_id, rua), não só "rua".
        cursor.execute("""
        DO $$
        BEGIN
            IF EXISTS (SELECT 1 FROM pg_constraint WHERE conname = 'notas_pkey') THEN
                ALTER TABLE notas DROP CONSTRAINT notas_pkey;
            END IF;
            IF NOT EXISTS (SELECT 1 FROM pg_constraint WHERE conname = 'notas_armazem_rua_key') THEN
                ALTER TABLE notas ADD CONSTRAINT notas_armazem_rua_key UNIQUE (armazem_id, rua);
            END IF;
        END $$;
        """)

        # sac_historico: (armazem_id, mes_ano) em vez de só "mes_ano".
        cursor.execute("""
        DO $$
        BEGIN
            IF EXISTS (SELECT 1 FROM pg_constraint WHERE conname = 'sac_historico_pkey') THEN
                ALTER TABLE sac_historico DROP CONSTRAINT sac_historico_pkey;
            END IF;
            IF NOT EXISTS (SELECT 1 FROM pg_constraint WHERE conname = 'sac_historico_armazem_mes_key') THEN
                ALTER TABLE sac_historico ADD CONSTRAINT sac_historico_armazem_mes_key UNIQUE (armazem_id, mes_ano);
            END IF;
        END $$;
        """)

        # rotativo_pessoas / rotativo_atividades: nome único por armazém,
        # não mais globalmente (dois armazéns podem ter uma pessoa "João").
        cursor.execute("""
        DO $$
        BEGIN
            IF EXISTS (SELECT 1 FROM pg_constraint WHERE conname = 'rotativo_pessoas_nome_key') THEN
                ALTER TABLE rotativo_pessoas DROP CONSTRAINT rotativo_pessoas_nome_key;
            END IF;
            IF NOT EXISTS (SELECT 1 FROM pg_constraint WHERE conname = 'rotativo_pessoas_armazem_nome_key') THEN
                ALTER TABLE rotativo_pessoas ADD CONSTRAINT rotativo_pessoas_armazem_nome_key UNIQUE (armazem_id, nome);
            END IF;
        END $$;
        """)

        cursor.execute("""
        DO $$
        BEGIN
            IF EXISTS (SELECT 1 FROM pg_constraint WHERE conname = 'rotativo_atividades_nome_key') THEN
                ALTER TABLE rotativo_atividades DROP CONSTRAINT rotativo_atividades_nome_key;
            END IF;
            IF NOT EXISTS (SELECT 1 FROM pg_constraint WHERE conname = 'rotativo_atividades_armazem_nome_key') THEN
                ALTER TABLE rotativo_atividades ADD CONSTRAINT rotativo_atividades_armazem_nome_key UNIQUE (armazem_id, nome);
            END IF;
        END $$;
        """)

        # ==============================
        # CRIA FUNDADOR
        # ==============================
        # Roda aqui (depois da migração multi-armazém) para já poder
        # gravar o armazem_id junto — se rodasse antes, em um banco que
        # já tivesse passado por essa migração, o INSERT quebraria por
        # não informar armazem_id (coluna NOT NULL).

        if USUARIO_FUNDADOR and SENHA_FUNDADOR:

            cursor.execute("""
            SELECT usuario
            FROM usuarios
            WHERE usuario = %s
            """, (
                USUARIO_FUNDADOR,
            ))

            fundador = cursor.fetchone()

            if not fundador:

                cursor.execute("""
                INSERT INTO usuarios
                    (
                    usuario,
                    senha,
                    tipo,
                    trocar_senha,
                    armazem_id
                )
                VALUES (%s, %s, %s, %s, %s)
                """, (
                    USUARIO_FUNDADOR,
                    SENHA_FUNDADOR,
                    "fundador",
                    0,
                    ID_ARMAZEM_PADRAO
                ))

        else:

            print(
                "⚠️ FUNDADOR_USUARIO / FUNDADOR_SENHA não configurados "
                "(secrets/variáveis de ambiente) — fundador não foi criado.",
                flush=True
            )

        # ==============================
        # MIGRAÇÃO: ÍNDICES DE PERFORMANCE
        # ==============================
        # armazem_id é FK, e o Postgres NÃO cria índice automático em
        # coluna de FK — só na chave primária. Como quase toda leitura
        # filtra por armazem_id (e várias ainda ordenam por data), sem
        # esses índices o banco varre a tabela inteira a cada consulta.
        # Isso não dói com poucos registros, mas piora conforme a base
        # cresce. IF NOT EXISTS torna seguro rodar em todo startup.

        cursor.execute("CREATE INDEX IF NOT EXISTS idx_historico_notas_rua_armazem ON historico_notas (armazem_id, rua, data_atualizacao DESC)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_historico_reman_armazem ON historico_remanejamento (armazem_id, data_hora DESC)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_analise_tecnica_armazem ON analise_tecnica (armazem_id, data_erro DESC)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_auditoria_armazem ON auditoria_atividades (armazem_id, data_atividade DESC)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_sac_historico_armazem ON sac_historico (armazem_id)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_remanejamento_armazem ON remanejamento (armazem_id)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_rotativo_pessoas_armazem ON rotativo_pessoas (armazem_id)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_rotativo_atividades_armazem ON rotativo_atividades (armazem_id)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_checklist_hidraulicos_armazem ON checklist_hidraulicos (armazem_id)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_checklist_carrinhos_armazem ON checklist_carrinhos (armazem_id)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_checklist_empilhadeiras_armazem ON checklist_empilhadeiras (armazem_id)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_checklist_pigmentacao_armazem ON checklist_pigmentacao (armazem_id)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_responsaveis_hidraulicos_armazem ON responsaveis_hidraulicos (armazem_id)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_responsaveis_carrinhos_armazem ON responsaveis_carrinhos (armazem_id)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_carrinhos_fixos_armazem ON carrinhos_fixos (armazem_id)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_usuarios_armazem ON usuarios (armazem_id)")

        conn.commit()

    finally:
        liberar(conn)

    return True


def inicializar_banco():

    _garantir_schema()

# ==================================================
# DASHBOARD
# ==================================================

@st.cache_data(ttl=30, show_spinner=False)
def listar_ruas(armazem_id):


    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        SELECT nome
        FROM ruas
        WHERE armazem_id = %s
        ORDER BY criado_em ASC, id ASC
        """, (armazem_id,))

        dados = [row[0] for row in cursor.fetchall()]

    finally:
        liberar(conn)

    return dados


def criar_rua(nome, armazem_id):

    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        INSERT INTO ruas (armazem_id, nome)
        VALUES (%s, %s)
        ON CONFLICT (armazem_id, nome) DO NOTHING
        """, (armazem_id, nome))

        conn.commit()

    finally:
        liberar(conn)

    listar_ruas.clear()


def excluir_rua(nome, armazem_id):

    conn = conectar()

    try:
        cursor = conn.cursor()

        # Remove o cadastro da rua e também os dados que já existiam
        # dela no Dashboard (nota atual + histórico), já que a rua
        # deixa de existir para esse armazém.
        cursor.execute(
            "DELETE FROM ruas WHERE armazem_id = %s AND nome = %s",
            (armazem_id, nome)
        )

        cursor.execute(
            "DELETE FROM notas WHERE armazem_id = %s AND rua = %s",
            (armazem_id, nome)
        )

        cursor.execute(
            "DELETE FROM historico_notas WHERE armazem_id = %s AND rua = %s",
            (armazem_id, nome)
        )

        conn.commit()

    finally:
        liberar(conn)

    listar_ruas.clear()
    ler_notas.clear()
    ler_duplas.clear()
    ler_tudo.clear()
    ler_historico_rua.clear()


# ==================================================
# PERFIS (nome, sobrenome, função, foto)
# ==================================================
# Resolve a ambiguidade de nomes repetidos (dois "João" etc.):
# cada usuário preenche seu próprio nome + sobrenome uma vez em
# "Perfil", e o app passa a reconhecer, em qualquer lugar que só
# tenha o primeiro nome digitado, se aquele texto bate com o nome,
# o sobrenome ou o nome completo do perfil de alguém.

import unicodedata
import base64


def _normalizar_texto(texto):

    texto = (texto or "").strip().lower()

    # remove acentos (ex.: "joão" -> "joao"), pra comparar sem
    # depender do usuário digitar acento igualzinho
    texto_sem_acento = unicodedata.normalize("NFKD", texto)

    return "".join(
        c for c in texto_sem_acento
        if not unicodedata.combining(c)
    )


def salvar_perfil(
    usuario,
    nome,
    sobrenome,
    funcao,
    foto,
    armazem_id
):

    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        INSERT INTO perfis
        (usuario, nome, sobrenome, funcao, foto, armazem_id, atualizado_em)
        VALUES (%s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP)
        ON CONFLICT (usuario) DO UPDATE SET
            nome = EXCLUDED.nome,
            sobrenome = EXCLUDED.sobrenome,
            funcao = EXCLUDED.funcao,
            foto = COALESCE(EXCLUDED.foto, perfis.foto),
            armazem_id = EXCLUDED.armazem_id,
            atualizado_em = CURRENT_TIMESTAMP
        """, (
            usuario,
            nome.strip(),
            sobrenome.strip(),
            (funcao or "").strip(),
            foto,
            armazem_id
        ))

        conn.commit()

    finally:
        liberar(conn)

    ler_perfil.clear()
    ler_perfis.clear()


@st.cache_data(ttl=30, show_spinner=False)
def ler_perfil(usuario):


    if not usuario:
        return None

    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        SELECT nome, sobrenome, funcao, foto, atualizado_em
        FROM perfis
        WHERE usuario = %s
        """, (usuario,))

        linha = cursor.fetchone()

    finally:
        liberar(conn)

    if not linha:
        return None

    return {
        "usuario": usuario,
        "nome": linha[0],
        "sobrenome": linha[1],
        "funcao": linha[2],
        "foto": linha[3],
        "atualizado_em": linha[4],
    }


@st.cache_data(ttl=30, show_spinner=False)
def ler_perfis(armazem_id):


    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        SELECT usuario, nome, sobrenome, funcao, foto
        FROM perfis
        WHERE armazem_id = %s
        ORDER BY nome, sobrenome
        """, (armazem_id,))

        linhas = cursor.fetchall()

    finally:
        liberar(conn)

    return [
        {
            "usuario": linha[0],
            "nome": linha[1],
            "sobrenome": linha[2],
            "funcao": linha[3],
            "foto": linha[4],
        }
        for linha in linhas
    ]


def nome_completo_perfil(perfil):

    if not perfil:
        return ""

    return f"{perfil['nome']} {perfil['sobrenome']}".strip()


def texto_bate_com_perfil(perfil, texto):
    """
    Compara um texto livre (ex.: um Nome digitado ao cadastrar um
    EPI, uma auditoria etc.) contra o perfil de alguém — bate se o
    texto for igual ao nome, ao sobrenome ou ao nome completo,
    sem diferenciar maiúsculas/minúsculas nem acentos.
    """

    if not perfil:
        return False

    texto_normalizado = _normalizar_texto(texto)

    if not texto_normalizado:
        return False

    candidatos = [
        perfil.get("nome"),
        perfil.get("sobrenome"),
        nome_completo_perfil(perfil),
    ]

    return any(
        _normalizar_texto(candidato) == texto_normalizado
        for candidato in candidatos
        if candidato
    )


def encontrar_perfil_por_nome(texto, armazem_id):
    """
    Dado um texto livre (ex.: "Paula" ou "Silva" ou "Paula Silva"),
    procura entre os perfis cadastrados do armazém alguém cujo nome,
    sobrenome ou nome completo bata. Se mais de uma pessoa bater
    (ex.: duas "Paula" diferentes), não arrisca escolher errado —
    devolve None nesse caso.
    """

    correspondentes = [
        perfil for perfil in ler_perfis(armazem_id)
        if texto_bate_com_perfil(perfil, texto)
    ]

    if len(correspondentes) == 1:
        return correspondentes[0]

    return None


def normalizar_nome_pessoa(texto, armazem_id):
    """
    Normaliza um nome digitado ou importado livremente (ex.: "alexandre",
    "Alexandre Vasques", "larissa santos"): se bater com o Perfil de
    alguém cadastrado nesse armazém (pelo nome, sobrenome ou nome
    completo, sem diferenciar maiúsculas/minúsculas nem acentos),
    devolve o nome completo do perfil (nome + sobrenome) — já
    corrigindo a capitalização e preenchendo o sobrenome que faltava.
    Se não bater com ninguém, apenas corrige a capitalização do texto
    digitado (Title Case).
    """

    texto = (texto or "").strip()

    if not texto:
        return texto

    perfil = encontrar_perfil_por_nome(texto, armazem_id)

    if perfil:
        return nome_completo_perfil(perfil)

    return texto.title()


def pessoa_pertence_ao_usuario(nome_pessoa, usuario_atual, armazem_id):
    """
    Verifica se um nome registrado livremente (análise técnica,
    auditoria etc.) corresponde ao usuário logado — seja pelo nome
    depois do prefixo tradicional (ex.: "Separador.Alexandre" bate
    com "Alexandre"), seja pelo Perfil dele (nome + sobrenome).
    Resolve o caso em que o registro tem o nome completo (ex.:
    "Alexandre Vasques") mas o login só tem o primeiro nome.
    """

    if not nome_pessoa:
        return False

    nome_normalizado = _normalizar_texto(nome_pessoa)

    if "." in (usuario_atual or ""):

        nome_login = usuario_atual.split(".", 1)[1].strip()

        if _normalizar_texto(nome_login) == nome_normalizado:
            return True

    perfil = ler_perfil(usuario_atual)

    return texto_bate_com_perfil(perfil, nome_pessoa)


def foto_para_base64(arquivo_upload, lado_max=256):
    """
    Recebe um arquivo enviado via st.file_uploader, redimensiona
    (lado máximo = lado_max) pra não pesar no banco, e devolve como
    string base64 (com prefixo data:image/...) pronta pra usar num
    <img src="...">.
    """

    from PIL import Image
    import io

    imagem = Image.open(arquivo_upload).convert("RGB")

    imagem.thumbnail((lado_max, lado_max))

    buffer = io.BytesIO()
    imagem.save(buffer, format="JPEG", quality=85)

    b64 = base64.b64encode(buffer.getvalue()).decode()

    return f"data:image/jpeg;base64,{b64}"


def chamados_ja_importados(armazem_id):
    """
    Conjunto com o número de todos os CHs (chamados) já registrados
    em Análise Técnica pra esse armazém — usado pra não duplicar um
    chamado que já foi importado de uma vez anterior da planilha.
    """

    return {
        registro["chamado"]
        for registro in ler_analise_tecnica(armazem_id)
        if registro.get("chamado")
    }


# ==================================================
# CONTROLE DE EPI's
# ==================================================
# Prefixos de usuário que podem receber a notificação de
# assinatura pendente para um determinado Nome (ex.: se o EPI
# foi registrado para "João", os usuários "Separador.João",
# "Conferente.João" e "Recebimento.João" — os que existirem —
# verão a pendência.

PREFIXOS_NOTIFICAVEIS_EPI = ["Separador.", "Conferente.", "Recebimento."]


def usuarios_alvo_epi(nome):

    nome_normalizado = nome.strip()

    return [
        f"{prefixo}{nome_normalizado}"
        for prefixo in PREFIXOS_NOTIFICAVEIS_EPI
    ]


def epi_pertence_ao_usuario(nome_epi, usuario_atual):
    """
    Compara sem diferenciar maiúsculas/minúsculas — "Separador.teste"
    (usuário) deve bater com um EPI cadastrado para "Teste", "TESTE"
    ou "teste". Também reconhece pelo Perfil (nome/sobrenome) do
    usuário logado, resolvendo o caso de nomes repetidos (dois
    "João" etc.) — se o EPI foi cadastrado como "João Ramires" e o
    usuário logado tem perfil com nome "João" e sobrenome "Ramires",
    bate certinho mesmo que outro "João" também exista.
    """

    usuario_normalizado = (usuario_atual or "").strip().lower()

    if any(
        candidato.strip().lower() == usuario_normalizado
        for candidato in usuarios_alvo_epi(nome_epi)
    ):
        return True

    perfil = ler_perfil(usuario_atual)

    return texto_bate_com_perfil(perfil, nome_epi)


@st.cache_data(ttl=30, show_spinner=False)
def ler_epis(armazem_id):


    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        SELECT id, nome, epi, data, assinatura, assinado_por,
            assinado_em, criado_por, criado_em
        FROM epis
        WHERE armazem_id = %s
        ORDER BY data DESC, id DESC
        """, (armazem_id,))

        colunas = [
            "id", "nome", "epi", "data", "assinatura", "assinado_por",
            "assinado_em", "criado_por", "criado_em"
        ]

        dados = [
            dict(zip(colunas, row))
            for row in cursor.fetchall()
        ]

    finally:
        liberar(conn)

    return dados


def criar_epi(nome, epi, data_epi, armazem_id, criado_por):

    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        INSERT INTO epis (armazem_id, nome, epi, data, criado_por)
        VALUES (%s, %s, %s, %s, %s)
        """, (armazem_id, nome, epi, data_epi, criado_por))

        conn.commit()

    finally:
        liberar(conn)

    ler_epis.clear()


def assinar_epi(id_epi, usuario, armazem_id):

    texto_assinatura = (
        f'Eu "{usuario}" assino declarando que recebi o item '
        f'conforme preenchido'
    )

    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        UPDATE epis
        SET assinatura = %s,
            assinado_por = %s,
            assinado_em = CURRENT_TIMESTAMP
        WHERE id = %s
        AND armazem_id = %s
        """, (texto_assinatura, usuario, id_epi, armazem_id))

        conn.commit()

    finally:
        liberar(conn)

    ler_epis.clear()


def excluir_epi(id_epi, armazem_id):

    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute(
            "DELETE FROM epis WHERE id = %s AND armazem_id = %s",
            (id_epi, armazem_id)
        )

        conn.commit()

    finally:
        liberar(conn)

    ler_epis.clear()


# ==================================================
# ALERTA: CHECKLIST DA SEXTA-FEIRA NÃO REALIZADO
# ==================================================
# Toda sexta-feira, cada responsável por hidráulico/carrinho
# (cadastrado em Equipamentos) deveria preencher o Checklist
# daquele equipamento. Esta função aponta quem ficou pendente
# na sexta-feira mais recente.

def _sexta_feira_mais_recente():

    hoje = date.today()

    # weekday(): segunda=0 ... sexta=4 ... domingo=6
    dias_desde_sexta = (hoje.weekday() - 4) % 7

    return hoje - timedelta(days=dias_desde_sexta)


def listar_pendentes_checklist_sexta(armazem_id):

    sexta_referencia = _sexta_feira_mais_recente()

    responsaveis_hidraulicos = ler_responsaveis_hidraulicos(armazem_id)
    responsaveis_carrinhos = ler_responsaveis_carrinhos(armazem_id)

    checklist_hidraulicos = ler_checklist_hidraulicos(armazem_id)
    checklist_carrinhos = ler_checklist_carrinhos(armazem_id)

    def _feito(nome, numero, registros):

        nome_normalizado = nome.strip().title()
        numero_normalizado = str(numero).strip()

        for registro in registros:

            if registro["data_checklist"] != sexta_referencia:
                continue

            if str(registro["numero"]).strip() != numero_normalizado:
                continue

            if registro["nome"].strip().title() != nome_normalizado:
                continue

            return True

        return False

    pendentes = []

    for item in responsaveis_hidraulicos:

        if not _feito(item["nome"], item["numero"], checklist_hidraulicos):

            pendentes.append({
                "nome": item["nome"],
                "numero": item["numero"],
                "tipo": "Hidráulico"
            })

    for item in responsaveis_carrinhos:

        if not _feito(item["nome"], item["numero"], checklist_carrinhos):

            pendentes.append({
                "nome": item["nome"],
                "numero": item["numero"],
                "tipo": "Carrinho"
            })

    return pendentes, sexta_referencia


# ==================================================
# TOP 3 DO MÊS FECHADO (RANKING DO DASHBOARD)
# ==================================================
# No último dia de cada mês, o Dashboard "fecha" o mês: aqui a
# gente reconstrói qual era a nota de cada rua naquele fechamento
# (a última nota registrada no histórico até aquela data) e monta
# o pódio das 3 melhores.

def _ultimo_dia_do_mes(ano, mes):

    if mes == 12:
        proximo_mes = date(ano + 1, 1, 1)
    else:
        proximo_mes = date(ano, mes + 1, 1)

    return proximo_mes - timedelta(days=1)


def _mes_fechado_mais_recente():

    # O servidor roda em UTC, então usar date.today() direto pode
    # considerar "hoje" um dia adiantado em relação ao horário real
    # de Campo Grande (ex.: 21h de 31/07 em Campo Grande já é
    # 01h de 01/08 em UTC). Por isso a data de "hoje" é sempre
    # calculada no fuso local antes de decidir qual mês fechou.
    hoje = datetime.now(ZoneInfo("America/Campo_Grande")).date()

    ultimo_dia_deste_mes = _ultimo_dia_do_mes(hoje.year, hoje.month)

    # Se hoje já é o último dia do mês, o fechamento é o de hoje.
    if hoje == ultimo_dia_deste_mes:
        return ultimo_dia_deste_mes

    # Senão, o fechamento é o do mês anterior.
    primeiro_dia_deste_mes = hoje.replace(day=1)

    return primeiro_dia_deste_mes - timedelta(days=1)


@st.cache_data(ttl=300, show_spinner=False)
def ler_top3_fechamento_mes(armazem_id):

    """
    Devolve (top3, data_fechamento): o pódio (até 3 ruas) com base
    na nota que cada rua tinha no fechamento do mês mais recente já
    encerrado (ou de hoje, se hoje for o último dia do mês).
    """

    data_fechamento = _mes_fechado_mais_recente()

    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        SELECT DISTINCT ON (rua)
            rua,
            nota,
            dupla
        FROM historico_notas
        WHERE armazem_id = %s
        AND (data_atualizacao AT TIME ZONE 'UTC' AT TIME ZONE 'America/Campo_Grande')::date <= %s
        ORDER BY rua, data_atualizacao DESC
        """, (
            armazem_id,
            data_fechamento
        ))

        dados = cursor.fetchall()

    finally:
        liberar(conn)

    ranking = [
        {
            "rua": linha[0],
            "nota": float(linha[1]) if linha[1] is not None else 0.0,
            "dupla": linha[2] or "Sem dupla"
        }
        for linha in dados
        if linha[1] is not None and linha[1] > 0
    ]

    ranking.sort(
        key=lambda item: item["nota"],
        reverse=True
    )

    return ranking[:3], data_fechamento


# ==================================================
# NOTIFICAÇÕES LIDAS (individual, por usuário)
# ==================================================
# Cada usuário tem seu próprio conjunto de notificações já lidas.
# Marcar uma como lida não afeta os outros usuários — cada um só
# deixa de ver a que ele mesmo marcou.

@st.cache_data(ttl=15, show_spinner=False)
def notificacoes_lidas_usuario(usuario, armazem_id):


    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        SELECT notificacao_id
        FROM notificacoes_lidas
        WHERE usuario = %s
        AND armazem_id = %s
        """, (
            usuario,
            armazem_id
        ))

        ids_lidas = {row[0] for row in cursor.fetchall()}

    finally:
        liberar(conn)

    return ids_lidas


def marcar_notificacao_lida(usuario, notificacao_id, armazem_id):

    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        INSERT INTO notificacoes_lidas
            (usuario, notificacao_id, armazem_id)
        VALUES (%s, %s, %s)
        ON CONFLICT (usuario, notificacao_id, armazem_id) DO NOTHING
        """, (
            usuario,
            notificacao_id,
            armazem_id
        ))

        conn.commit()

    finally:
        liberar(conn)

    notificacoes_lidas_usuario.clear()


# ==================================================
# NOTIFICAÇÕES EXCLUÍDAS (global — Gestão/Fundador)
# ==================================================
# Ao excluir, a notificação some da central de notificações para
# TODOS os usuários daquele armazém (diferente de "lida", que é
# só individual).

@st.cache_data(ttl=15, show_spinner=False)
def notificacoes_excluidas(armazem_id):


    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        SELECT notificacao_id
        FROM notificacoes_excluidas
        WHERE armazem_id = %s
        """, (
            armazem_id,
        ))

        ids_excluidas = {row[0] for row in cursor.fetchall()}

    finally:
        liberar(conn)

    return ids_excluidas


def excluir_notificacao(notificacao_id, armazem_id, usuario):

    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        INSERT INTO notificacoes_excluidas
            (notificacao_id, armazem_id, excluida_por)
        VALUES (%s, %s, %s)
        ON CONFLICT (notificacao_id, armazem_id) DO NOTHING
        """, (
            notificacao_id,
            armazem_id,
            usuario
        ))

        conn.commit()

    finally:
        liberar(conn)

    notificacoes_excluidas.clear()


@st.cache_data(ttl=30, show_spinner=False)
def ler_notas(armazem_id):


    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        SELECT
            rua,
            nota
        FROM notas
        WHERE armazem_id = %s
        """, (armazem_id,))

        dados = {
            row[0]: float(row[1]) if row[1] is not None else 0
            for row in cursor.fetchall()
        }

    finally:
        liberar(conn)

    return dados


@st.cache_data(ttl=30, show_spinner=False)
def ler_duplas(armazem_id):


    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        SELECT
            rua,
            dupla
        FROM notas
        WHERE armazem_id = %s
        """, (armazem_id,))

        dados = {
            row[0]: row[1]
            for row in cursor.fetchall()
        }

    finally:
        liberar(conn)

    return dados


@st.cache_data(ttl=30, show_spinner=False)
def ler_tudo(armazem_id):


    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        SELECT
            rua,
            nota,
            dupla
        FROM notas
        WHERE armazem_id = %s
        """, (armazem_id,))

        dados = {}

        for row in cursor.fetchall():

            dados[row[0]] = {
                "nota": float(row[1]) if row[1] is not None else 0,
                "dupla": row[2]
            }

    finally:
        liberar(conn)

    return dados


def salvar_dados(
    rua,
    nota,
    dupla,
    armazem_id,
    usuario=None
):

    conn = conectar()

    try:
        cursor = conn.cursor()

        # Atualiza o painel atual
        cursor.execute("""
        INSERT INTO notas
        (
            rua,
            nota,
            dupla,
            atualizado_por,
            armazem_id
        )
        VALUES (%s, %s, %s, %s, %s)

        ON CONFLICT (armazem_id, rua)
        DO UPDATE SET
            nota = EXCLUDED.nota,
            dupla = EXCLUDED.dupla,
            atualizado_por = EXCLUDED.atualizado_por
        """, (
            rua,
            nota,
            dupla,
            usuario,
            armazem_id
        ))

        # Salva histórico
        cursor.execute("""
        INSERT INTO historico_notas
        (
            rua,
            nota,
            dupla,
            usuario,
            armazem_id
        )
        VALUES (%s, %s, %s, %s, %s)
        """, (
            rua,
            nota,
            dupla,
            usuario,
            armazem_id
        ))

        conn.commit()

    finally:
        liberar(conn)

    # limpa o cache para refletir o dado novo imediatamente
    ler_notas.clear()
    ler_duplas.clear()
    ler_tudo.clear()
    ler_historico_rua.clear()


@st.cache_data(ttl=30, show_spinner=False)
def ler_historico_rua(
    rua,
    armazem_id,
    limite=10
):


    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        SELECT
            nota,
            dupla,
            data_atualizacao,
            usuario
        FROM historico_notas
        WHERE rua = %s
        AND armazem_id = %s
        ORDER BY data_atualizacao DESC
        LIMIT %s
        """, (
            rua,
            armazem_id,
            limite
        ))

        dados = cursor.fetchall()

    finally:
        liberar(conn)

    return dados

# ==================================================
# REMANEJAMENTO
# ==================================================

@st.cache_data(ttl=30, show_spinner=False)
def ler_remanejamentos(armazem_id):


    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        SELECT
            id,
            item,
            prioridade,
            criado_por
        FROM remanejamento
        WHERE armazem_id = %s
        ORDER BY
            CASE prioridade
                WHEN 'Alta' THEN 1
                WHEN 'Média' THEN 2
                ELSE 3
            END,
            id DESC
        """, (armazem_id,))

        dados = []

        for row in cursor.fetchall():

            dados.append({
                "id": row[0],
                "nome": row[1],
                "prioridade": row[2],
                "criado_por": row[3]
            })

    finally:
        liberar(conn)

    return dados


def adicionar_remanejamento(
    item,
    armazem_id,
    prioridade="Normal",
    usuario=None
):

    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        INSERT INTO remanejamento
        (
            item,
            prioridade,
            criado_por,
            armazem_id
        )
        VALUES (%s, %s, %s, %s)
        """, (
            item,
            prioridade,
            usuario,
            armazem_id
        ))

        cursor.execute("""
        INSERT INTO historico_remanejamento
        (
            item,
            prioridade,
            usuario,
            armazem_id
        )
        VALUES (%s, %s, %s, %s)
        """, (
            item,
            prioridade,
            usuario,
            armazem_id
        ))

        conn.commit()

    finally:
        liberar(conn)

    ler_remanejamentos.clear()
    ler_historico_remanejamento.clear()
    total_remanejamentos.clear()


def excluir_remanejamento(
    id_item,
    armazem_id
):

    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        DELETE FROM remanejamento
        WHERE id = %s
        AND armazem_id = %s
        """, (
            id_item,
            armazem_id
        ))

        conn.commit()

    finally:
        liberar(conn)

    ler_remanejamentos.clear()
    total_remanejamentos.clear()


def excluir_remanejamento_lote(
    ids_itens,
    armazem_id
):

    if not ids_itens:
        return

    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        DELETE FROM remanejamento
        WHERE id = ANY(%s)
        AND armazem_id = %s
        """, (
            ids_itens,
            armazem_id
        ))

        conn.commit()

    finally:
        liberar(conn)

    ler_remanejamentos.clear()
    total_remanejamentos.clear()


@st.cache_data(ttl=30, show_spinner=False)
def total_remanejamentos(armazem_id):


    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        SELECT COUNT(*)
        FROM remanejamento
        WHERE armazem_id = %s
        """, (armazem_id,))

        total = cursor.fetchone()[0]

    finally:
        liberar(conn)

    return total

@st.cache_data(ttl=30, show_spinner=False)
def ler_historico_remanejamento(
    armazem_id,
    limite=20
):


    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        SELECT
            item,
            prioridade,
            data_hora,
            usuario
        FROM historico_remanejamento
        WHERE armazem_id = %s
        ORDER BY data_hora DESC
        LIMIT %s
        """, (armazem_id, limite))

        dados = cursor.fetchall()

    finally:
        liberar(conn)

    return dados


# ==================================================
# REMANEJAMENTO AGENDADO (por horário/dia da semana)
# ==================================================

@st.cache_data(ttl=30, show_spinner=False)
def ler_remanejamentos_agendados(armazem_id):


    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        SELECT
            id,
            item,
            prioridade,
            hora_inicio,
            hora_fim,
            dias_semana,
            criado_por
        FROM remanejamento_agendado
        WHERE armazem_id = %s
        ORDER BY hora_inicio, item
        """, (armazem_id,))

        dados = []

        for row in cursor.fetchall():

            dados.append({
                "id": row[0],
                "nome": row[1],
                "prioridade": row[2],
                "hora_inicio": row[3],
                "hora_fim": row[4],
                "dias_semana": row[5] or [],
                "criado_por": row[6]
            })

    finally:
        liberar(conn)

    return dados


def criar_remanejamento_agendado(
    item,
    prioridade,
    hora_inicio,
    hora_fim,
    dias_semana,
    armazem_id,
    usuario=None
):

    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        INSERT INTO remanejamento_agendado
        (
            item,
            prioridade,
            hora_inicio,
            hora_fim,
            dias_semana,
            criado_por,
            armazem_id
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (
            item,
            prioridade,
            hora_inicio,
            hora_fim,
            dias_semana,
            usuario,
            armazem_id
        ))

        conn.commit()

    finally:
        liberar(conn)

    ler_remanejamentos_agendados.clear()


def excluir_remanejamento_agendado(
    id_item,
    armazem_id
):

    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        DELETE FROM remanejamento_agendado
        WHERE id = %s
        AND armazem_id = %s
        """, (
            id_item,
            armazem_id
        ))

        conn.commit()

    finally:
        liberar(conn)

    ler_remanejamentos_agendados.clear()


# ==================================================
# SAC
# ==================================================

def atualizar_sac_mensal(
    reclamacoes,
    meta,
    armazem_id,
    usuario=None
):

    mes_ano = datetime.now().strftime(
        "%Y-%m"
    )

    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        INSERT INTO sac_historico
        (
            mes_ano,
            reclamacoes,
            meta,
            atualizado_por,
            atualizado_em,
            armazem_id
        )
        VALUES (%s, %s, %s, %s, CURRENT_TIMESTAMP, %s)

        ON CONFLICT (armazem_id, mes_ano)
        DO UPDATE SET
            reclamacoes = EXCLUDED.reclamacoes,
            meta = EXCLUDED.meta,
            atualizado_por = EXCLUDED.atualizado_por,
            atualizado_em = CURRENT_TIMESTAMP
        """, (
            mes_ano,
            reclamacoes,
            meta,
            usuario,
            armazem_id
        ))

        conn.commit()

    finally:
        liberar(conn)

    ler_historico_sac.clear()
    total_reclamacoes.clear()


@st.cache_data(ttl=30, show_spinner=False)
def ler_historico_sac(armazem_id):


    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        SELECT
            mes_ano,
            reclamacoes,
            meta,
            atualizado_por,
            atualizado_em
        FROM sac_historico
        WHERE armazem_id = %s
        ORDER BY mes_ano ASC
        """, (armazem_id,))

        dados = cursor.fetchall()

    finally:
        liberar(conn)

    return dados


@st.cache_data(ttl=30, show_spinner=False)
def total_reclamacoes(armazem_id):


    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        SELECT
            COALESCE(
                SUM(reclamacoes),
                0
            )
        FROM sac_historico
        WHERE armazem_id = %s
        """, (armazem_id,))

        total = cursor.fetchone()[0]

    finally:
        liberar(conn)

    return total

# ==================================================
# ANÁLISE TÉCNICA (SAC)
# ==================================================

# ==================================================
# PLANILHA VINCULADA DO SAC (importação automática)
# ==================================================
# A pessoa envia (upload) o Excel pelo Administrativo. O app guarda
# essa cópia no banco e, periodicamente, relê a aba do MÊS ATUAL
# (ex.: "Agosto 2026") procurando linhas com "CH" (chamado) que
# ainda não existem em analise_tecnica — cada linha nova vira um
# chamado automaticamente, com o Separador/Conferente já notificados
# do mesmo jeito que um cadastro manual notificaria.

MESES_PT = [
    "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
    "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"
]


def nome_aba_mes_atual():

    agora = datetime.now(ZoneInfo("America/Campo_Grande"))

    return f"{MESES_PT[agora.month - 1]} {agora.year}"


def salvar_planilha_sac(armazem_id, nome_arquivo, conteudo_bytes, usuario):

    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        INSERT INTO planilha_sac
        (armazem_id, nome_arquivo, conteudo, enviado_por, enviado_em, ultima_checagem, ultimo_resultado)
        VALUES (%s, %s, %s, %s, CURRENT_TIMESTAMP, NULL, NULL)
        ON CONFLICT (armazem_id) DO UPDATE SET
            nome_arquivo = EXCLUDED.nome_arquivo,
            conteudo = EXCLUDED.conteudo,
            enviado_por = EXCLUDED.enviado_por,
            enviado_em = CURRENT_TIMESTAMP,
            ultima_checagem = NULL,
            ultimo_resultado = NULL
        """, (
            armazem_id,
            nome_arquivo,
            psycopg2.Binary(conteudo_bytes),
            usuario
        ))

        conn.commit()

    finally:
        liberar(conn)

    ler_planilha_sac.clear()


@st.cache_data(ttl=30, show_spinner=False)
def ler_planilha_sac(armazem_id):


    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        SELECT nome_arquivo, conteudo, enviado_por, enviado_em,
               ultima_checagem, ultimo_resultado
        FROM planilha_sac
        WHERE armazem_id = %s
        """, (armazem_id,))

        linha = cursor.fetchone()

    finally:
        liberar(conn)

    if not linha:
        return None

    return {
        "nome_arquivo": linha[0],
        "conteudo": bytes(linha[1]) if linha[1] is not None else None,
        "enviado_por": linha[2],
        "enviado_em": linha[3],
        "ultima_checagem": linha[4],
        "ultimo_resultado": linha[5],
    }


def _registrar_checagem_planilha_sac(armazem_id, resultado_texto):

    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        UPDATE planilha_sac
        SET ultima_checagem = CURRENT_TIMESTAMP,
            ultimo_resultado = %s
        WHERE armazem_id = %s
        """, (resultado_texto, armazem_id))

        conn.commit()

    finally:
        liberar(conn)

    ler_planilha_sac.clear()


def _valor_texto(valor):

    if valor is None:
        return ""

    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))

    return str(valor).strip()


def _valor_hora(valor):

    from datetime import time as time_type, datetime as datetime_type

    if valor is None:
        return None

    if isinstance(valor, time_type):
        return valor

    if isinstance(valor, datetime_type):
        return valor.time()

    texto = str(valor).strip()

    for formato in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime_type.strptime(texto, formato).time()
        except ValueError:
            continue

    return None


def _valor_data(valor):

    if valor is None:
        return None

    if hasattr(valor, "date") and callable(valor.date):
        return valor.date()

    if hasattr(valor, "year"):
        return valor

    return None


def importar_planilha_sac(armazem_id, usuario="Planilha Vinculada"):
    """
    Relê a cópia da planilha guardada para este armazém, acha a aba
    do mês atual e cria em analise_tecnica os chamados ("CH") que
    ainda não existem — comparando pelo número do chamado. Devolve
    um dicionário com o resultado (pra mostrar na tela).
    """

    import openpyxl
    import io as io_module

    registro_planilha = ler_planilha_sac(armazem_id)

    if not registro_planilha or not registro_planilha.get("conteudo"):
        return {"ok": False, "mensagem": "Nenhuma planilha vinculada ainda."}

    aba_alvo = nome_aba_mes_atual()

    try:
        pasta_trabalho = openpyxl.load_workbook(
            io_module.BytesIO(registro_planilha["conteudo"]),
            data_only=True
        )
    except Exception as erro:
        resultado = f"Erro ao abrir a planilha: {erro}"
        _registrar_checagem_planilha_sac(armazem_id, resultado)
        return {"ok": False, "mensagem": resultado}

    if aba_alvo not in pasta_trabalho.sheetnames:
        resultado = f"Aba '{aba_alvo}' não encontrada na planilha."
        _registrar_checagem_planilha_sac(armazem_id, resultado)
        return {"ok": False, "mensagem": resultado}

    planilha = pasta_trabalho[aba_alvo]

    chamados_existentes = {
        (registro.get("chamado") or "").strip().lower()
        for registro in ler_analise_tecnica(armazem_id)
    }

    novos = 0
    ignorados = 0

    linhas = planilha.iter_rows(min_row=2, values_only=True)

    for linha in linhas:

        if linha is None or len(linha) < 19:
            continue

        (
            data_col, ch_col, cliente_col, nf_col, descricao_col,
            cod_pdt_col, produto_col, tratativa_col, data2_col, hora_col,
            separador_col, volume_col, carga_col, regiao_col, motorista_col,
            balanca_col, conf_col, tipo_col, _separador2_col
        ) = linha[:19]

        chamado = _valor_texto(ch_col)

        if not chamado:
            continue

        if chamado.strip().lower() in chamados_existentes:
            ignorados += 1
            continue

        data_erro = _valor_data(data_col) or datetime.now(ZoneInfo("America/Campo_Grande")).date()

        separador = _valor_texto(separador_col) or None

        conferente_bruto = _valor_texto(conf_col)

        conferente = (
            conferente_bruto
            if conferente_bruto and conferente_bruto.lower() not in ("não", "nao", "-")
            else None
        )

        dados = {
            "tipo_erro": _valor_texto(tipo_col) or "Não informado",
            "data_erro": data_erro,
            "data_fechamento": None,
            "descricao": _valor_texto(descricao_col),
            "chamado": chamado,
            "cliente": _valor_texto(cliente_col),
            "nota_fiscal": _valor_texto(nf_col),
            "cod_produto": _valor_texto(cod_pdt_col),
            "produto": _valor_texto(produto_col),
            "tratativa": _valor_texto(tratativa_col) or "Não informado",
            "hora": _valor_hora(hora_col),
            "separador": separador,
            "volume": _valor_texto(volume_col),
            "carga": _valor_texto(carga_col),
            "regiao": _valor_texto(regiao_col),
            "motorista": _valor_texto(motorista_col),
            "balanca": _valor_texto(balanca_col),
            "conferente": conferente,
        }

        vinculos = []

        if separador:
            vinculos.append({"nome": separador, "papel": "Separador"})

        if conferente:
            vinculos.append({"nome": conferente, "papel": "Conferente"})

        adicionar_analise_tecnica(dados, vinculos, armazem_id, usuario=usuario)

        chamados_existentes.add(chamado.strip().lower())

        novos += 1

    if novos:
        resultado = f"{novos} chamado(s) novo(s) importado(s) de '{aba_alvo}'."
    else:
        resultado = f"Nenhum chamado novo em '{aba_alvo}' ({ignorados} já existente(s))."

    _registrar_checagem_planilha_sac(armazem_id, resultado)

    return {
        "ok": True,
        "mensagem": resultado,
        "novos": novos,
        "ignorados": ignorados,
        "aba": aba_alvo,
    }


def adicionar_analise_tecnica(
    dados,
    vinculos,
    armazem_id,
    usuario=None
):

    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        INSERT INTO analise_tecnica
        (
            nome, tipo_erro, data_erro, data_fechamento, descricao,
            chamado, cliente, nota_fiscal, cod_produto, produto,
            tratativa, hora, separador, volume, carga, regiao,
            motorista, balanca, conferente, vinculos_notificados,
            registrado_por, armazem_id
        )
        VALUES (
            %s, %s, %s, %s, %s,
            %s, %s, %s, %s, %s,
            %s, %s, %s, %s, %s, %s,
            %s, %s, %s, %s,
            %s, %s
        )
        """, (
            None, dados["tipo_erro"], dados["data_erro"], dados.get("data_fechamento"), dados["descricao"],
            dados["chamado"], dados["cliente"], dados["nota_fiscal"], dados["cod_produto"], dados["produto"],
            dados["tratativa"], dados["hora"], dados["separador"], dados["volume"], dados["carga"], dados["regiao"],
            dados["motorista"], dados["balanca"], dados["conferente"], Json(vinculos),
            usuario, armazem_id
        ))

        conn.commit()

    finally:
        liberar(conn)

    ler_analise_tecnica.clear()


@st.cache_data(ttl=30, show_spinner=False)
def ler_analise_tecnica(armazem_id):


    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        SELECT
            id, nome, tipo_erro, data_erro, data_fechamento, descricao,
            chamado, cliente, nota_fiscal, cod_produto, produto,
            tratativa, hora, separador, volume, carga, regiao,
            motorista, balanca, conferente, vinculos_notificados,
            registrado_por
        FROM analise_tecnica
        WHERE armazem_id = %s
        ORDER BY data_erro DESC
        """, (armazem_id,))

        colunas = [
            "id", "nome", "tipo_erro", "data_erro", "data_fechamento", "descricao",
            "chamado", "cliente", "nota_fiscal", "cod_produto", "produto",
            "tratativa", "hora", "separador", "volume", "carga", "regiao",
            "motorista", "balanca", "conferente", "vinculos_notificados",
            "registrado_por"
        ]

        dados = []

        for row in cursor.fetchall():

            dados.append(
                dict(zip(colunas, row))
            )

    finally:
        liberar(conn)

    return dados


def finalizar_analise_tecnica(
    id_registro,
    data_fechamento,
    armazem_id
):

    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        UPDATE analise_tecnica
        SET data_fechamento = %s
        WHERE id = %s
        AND armazem_id = %s
        """, (
            data_fechamento,
            id_registro,
            armazem_id
        ))

        conn.commit()

    finally:
        liberar(conn)

    ler_analise_tecnica.clear()


def excluir_analise_tecnica(
    id_registro,
    armazem_id
):

    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        DELETE FROM analise_tecnica
        WHERE id = %s
        AND armazem_id = %s
        """, (
            id_registro,
            armazem_id
        ))

        conn.commit()

    finally:
        liberar(conn)

    ler_analise_tecnica.clear()


def excluir_analise_tecnica_lote(
    ids_registros,
    armazem_id
):

    if not ids_registros:
        return

    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        DELETE FROM analise_tecnica
        WHERE id = ANY(%s)
        AND armazem_id = %s
        """, (
            ids_registros,
            armazem_id
        ))

        conn.commit()

    finally:
        liberar(conn)

    ler_analise_tecnica.clear()

# ==================================================
# AUDITORIA DE ATIVIDADES
# ==================================================


def adicionar_auditoria(
    nome,
    funcao,
    qtd_acertos,
    qtd_erros,
    data_atividade,
    descricao,
    armazem_id,
    usuario=None
):

    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        INSERT INTO auditoria_atividades
        (
            nome, funcao, qtd_acertos, qtd_erros,
            data_atividade, descricao, registrado_por, armazem_id
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            nome, funcao, qtd_acertos, qtd_erros,
            data_atividade, descricao, usuario, armazem_id
        ))

        conn.commit()

    finally:
        liberar(conn)

    ler_auditoria.clear()


@st.cache_data(ttl=30, show_spinner=False)
def ler_auditoria(armazem_id):


    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        SELECT
            id, nome, funcao, qtd_acertos, qtd_erros,
            data_atividade, descricao, registrado_por
        FROM auditoria_atividades
        WHERE armazem_id = %s
        ORDER BY data_atividade DESC
        """, (armazem_id,))

        colunas = [
            "id", "nome", "funcao", "qtd_acertos", "qtd_erros",
            "data_atividade", "descricao", "registrado_por"
        ]

        dados = []

        for row in cursor.fetchall():

            dados.append(
                dict(zip(colunas, row))
            )

    finally:
        liberar(conn)

    return dados


def excluir_auditoria(
    id_registro,
    armazem_id
):

    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        DELETE FROM auditoria_atividades
        WHERE id = %s
        AND armazem_id = %s
        """, (
            id_registro,
            armazem_id
        ))

        conn.commit()

    finally:
        liberar(conn)

    ler_auditoria.clear()


def excluir_auditoria_lote(
    ids_registros,
    armazem_id
):

    if not ids_registros:
        return

    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        DELETE FROM auditoria_atividades
        WHERE id = ANY(%s)
        AND armazem_id = %s
        """, (
            ids_registros,
            armazem_id
        ))

        conn.commit()

    finally:
        liberar(conn)

    ler_auditoria.clear()

# ==================================================
# RODÍZIO - FIM DE EXPEDIENTE
# ==================================================


def adicionar_pessoa_rotativo(nome, armazem_id):

    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        INSERT INTO rotativo_pessoas (nome, armazem_id)
        VALUES (%s, %s)
        ON CONFLICT (armazem_id, nome) DO NOTHING
        """, (
            nome,
            armazem_id
        ))

        conn.commit()

    finally:
        liberar(conn)

    listar_pessoas_rotativo.clear()


@st.cache_data(ttl=30, show_spinner=False)
def listar_pessoas_rotativo(armazem_id):


    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        SELECT id, nome
        FROM rotativo_pessoas
        WHERE armazem_id = %s
        ORDER BY id
        """, (armazem_id,))

        dados = cursor.fetchall()

    finally:
        liberar(conn)

    return dados


def excluir_pessoa_rotativo(id_pessoa, armazem_id):

    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        DELETE FROM rotativo_pessoas
        WHERE id = %s
        AND armazem_id = %s
        """, (
            id_pessoa,
            armazem_id
        ))

        conn.commit()

    finally:
        liberar(conn)

    listar_pessoas_rotativo.clear()


def adicionar_atividade_rotativo(
    nome,
    armazem_id,
    tipo="rotativo",
    pessoa_fixa=None
):

    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        INSERT INTO rotativo_atividades (nome, tipo, pessoa_fixa, armazem_id)
        VALUES (%s, %s, %s, %s)
        ON CONFLICT (armazem_id, nome) DO NOTHING
        """, (
            nome, tipo, pessoa_fixa, armazem_id
        ))

        conn.commit()

    finally:
        liberar(conn)

    listar_atividades_rotativo.clear()


@st.cache_data(ttl=30, show_spinner=False)
def listar_atividades_rotativo(armazem_id):


    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        SELECT id, nome, tipo, pessoa_fixa
        FROM rotativo_atividades
        WHERE armazem_id = %s
        ORDER BY id
        """, (armazem_id,))

        dados = cursor.fetchall()

    finally:
        liberar(conn)

    return dados


def excluir_atividade_rotativo(id_atividade, armazem_id):

    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        DELETE FROM rotativo_atividades
        WHERE id = %s
        AND armazem_id = %s
        """, (
            id_atividade,
            armazem_id
        ))

        conn.commit()

    finally:
        liberar(conn)

    listar_atividades_rotativo.clear()

# ==================================================
# CHECKLISTS
# ==================================================


def _adicionar_checklist(tabela, nome, numero, data_checklist, status, descricao, armazem_id):

    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute(f"""
        INSERT INTO {tabela}
        (nome, numero, data_checklist, status, descricao, armazem_id)
        VALUES (%s, %s, %s, %s, %s, %s)
        """, (
            nome, numero, data_checklist, status, descricao, armazem_id
        ))

        conn.commit()

    finally:
        liberar(conn)


def _ler_checklist(tabela, armazem_id):

    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute(f"""
        SELECT id, nome, numero, data_checklist, status, descricao,
            em_manutencao, manutencao_motivo, manutencao_enviado_por,
            manutencao_enviado_em, manutencao_retornado_por, manutencao_retornado_em
        FROM {tabela}
        WHERE armazem_id = %s
        ORDER BY data_checklist DESC, id DESC
        """, (armazem_id,))

        colunas = [
            "id", "nome", "numero", "data_checklist", "status", "descricao",
            "em_manutencao", "manutencao_motivo", "manutencao_enviado_por",
            "manutencao_enviado_em", "manutencao_retornado_por", "manutencao_retornado_em"
        ]

        dados = [
            dict(zip(colunas, row))
            for row in cursor.fetchall()
        ]

    finally:
        liberar(conn)

    return dados


def _editar_checklist(tabela, id_registro, nome, numero, data_checklist, status, descricao, armazem_id):

    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute(f"""
        UPDATE {tabela}
        SET nome = %s,
            numero = %s,
            data_checklist = %s,
            status = %s,
            descricao = %s
        WHERE id = %s
        AND armazem_id = %s
        """, (
            nome, numero, data_checklist, status, descricao, id_registro, armazem_id
        ))

        conn.commit()

    finally:
        liberar(conn)


def _enviar_manutencao_checklist(tabela, id_registro, motivo, usuario, armazem_id):

    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute(f"""
        UPDATE {tabela}
        SET em_manutencao = TRUE,
            manutencao_motivo = %s,
            manutencao_enviado_por = %s,
            manutencao_enviado_em = CURRENT_TIMESTAMP,
            manutencao_retornado_por = NULL,
            manutencao_retornado_em = NULL
        WHERE id = %s
        AND armazem_id = %s
        """, (
            motivo, usuario, id_registro, armazem_id
        ))

        conn.commit()

    finally:
        liberar(conn)


def _retornar_manutencao_checklist(tabela, id_registro, usuario, armazem_id):

    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute(f"""
        UPDATE {tabela}
        SET em_manutencao = FALSE,
            manutencao_retornado_por = %s,
            manutencao_retornado_em = CURRENT_TIMESTAMP
        WHERE id = %s
        AND armazem_id = %s
        """, (
            usuario, id_registro, armazem_id
        ))

        conn.commit()

    finally:
        liberar(conn)


def _excluir_checklist(tabela, id_registro, armazem_id):

    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute(f"""
        DELETE FROM {tabela}
        WHERE id = %s
        AND armazem_id = %s
        """, (
            id_registro, armazem_id
        ))

        conn.commit()

    finally:
        liberar(conn)


def adicionar_checklist_hidraulico(nome, numero, data_checklist, status, descricao, armazem_id):

    _adicionar_checklist(
        "checklist_hidraulicos",
        nome, numero, data_checklist, status, descricao, armazem_id
    )

    ler_checklist_hidraulicos.clear()


@st.cache_data(ttl=30, show_spinner=False)
def ler_checklist_hidraulicos(armazem_id):


    return _ler_checklist("checklist_hidraulicos", armazem_id)


def editar_checklist_hidraulico(id_registro, nome, numero, data_checklist, status, descricao, armazem_id):

    _editar_checklist(
        "checklist_hidraulicos",
        id_registro, nome, numero, data_checklist, status, descricao, armazem_id
    )

    ler_checklist_hidraulicos.clear()


def enviar_manutencao_hidraulico(id_registro, motivo, usuario, armazem_id):

    _enviar_manutencao_checklist(
        "checklist_hidraulicos", id_registro, motivo, usuario, armazem_id
    )

    ler_checklist_hidraulicos.clear()


def retornar_manutencao_hidraulico(id_registro, usuario, armazem_id):

    _retornar_manutencao_checklist(
        "checklist_hidraulicos", id_registro, usuario, armazem_id
    )

    ler_checklist_hidraulicos.clear()


def excluir_checklist_hidraulico(id_registro, armazem_id):

    _excluir_checklist(
        "checklist_hidraulicos", id_registro, armazem_id
    )

    ler_checklist_hidraulicos.clear()


def adicionar_checklist_carrinho(nome, numero, data_checklist, status, descricao, armazem_id):

    _adicionar_checklist(
        "checklist_carrinhos",
        nome, numero, data_checklist, status, descricao, armazem_id
    )

    ler_checklist_carrinhos.clear()


@st.cache_data(ttl=30, show_spinner=False)
def ler_checklist_carrinhos(armazem_id):


    return _ler_checklist("checklist_carrinhos", armazem_id)


def editar_checklist_carrinho(id_registro, nome, numero, data_checklist, status, descricao, armazem_id):

    _editar_checklist(
        "checklist_carrinhos",
        id_registro, nome, numero, data_checklist, status, descricao, armazem_id
    )

    ler_checklist_carrinhos.clear()


def enviar_manutencao_carrinho(id_registro, motivo, usuario, armazem_id):

    _enviar_manutencao_checklist(
        "checklist_carrinhos", id_registro, motivo, usuario, armazem_id
    )

    ler_checklist_carrinhos.clear()


def retornar_manutencao_carrinho(id_registro, usuario, armazem_id):

    _retornar_manutencao_checklist(
        "checklist_carrinhos", id_registro, usuario, armazem_id
    )

    ler_checklist_carrinhos.clear()


def excluir_checklist_carrinho(id_registro, armazem_id):

    _excluir_checklist(
        "checklist_carrinhos", id_registro, armazem_id
    )

    ler_checklist_carrinhos.clear()


def adicionar_checklist_empilhadeira(nome, numero, data_checklist, status, descricao, armazem_id):

    _adicionar_checklist(
        "checklist_empilhadeiras",
        nome, numero, data_checklist, status, descricao, armazem_id
    )

    ler_checklist_empilhadeiras.clear()


@st.cache_data(ttl=30, show_spinner=False)
def ler_checklist_empilhadeiras(armazem_id):


    return _ler_checklist("checklist_empilhadeiras", armazem_id)


def editar_checklist_empilhadeira(id_registro, nome, numero, data_checklist, status, descricao, armazem_id):

    _editar_checklist(
        "checklist_empilhadeiras",
        id_registro, nome, numero, data_checklist, status, descricao, armazem_id
    )

    ler_checklist_empilhadeiras.clear()


def enviar_manutencao_empilhadeira(id_registro, motivo, usuario, armazem_id):

    _enviar_manutencao_checklist(
        "checklist_empilhadeiras", id_registro, motivo, usuario, armazem_id
    )

    ler_checklist_empilhadeiras.clear()


def retornar_manutencao_empilhadeira(id_registro, usuario, armazem_id):

    _retornar_manutencao_checklist(
        "checklist_empilhadeiras", id_registro, usuario, armazem_id
    )

    ler_checklist_empilhadeiras.clear()


def excluir_checklist_empilhadeira(id_registro, armazem_id):

    _excluir_checklist(
        "checklist_empilhadeiras", id_registro, armazem_id
    )

    ler_checklist_empilhadeiras.clear()


def adicionar_checklist_pigmentacao(nome, data_checklist, status, descricao, armazem_id):

    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        INSERT INTO checklist_pigmentacao
        (nome, data_checklist, status, descricao, armazem_id)
        VALUES (%s, %s, %s, %s, %s)
        """, (
            nome, data_checklist, status, descricao, armazem_id
        ))

        conn.commit()

    finally:
        liberar(conn)

    ler_checklist_pigmentacao.clear()


@st.cache_data(ttl=30, show_spinner=False)
def ler_checklist_pigmentacao(armazem_id):


    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        SELECT id, nome, data_checklist, status, descricao
        FROM checklist_pigmentacao
        WHERE armazem_id = %s
        ORDER BY data_checklist DESC, id DESC
        """, (armazem_id,))

        colunas = ["id", "nome", "data_checklist", "status", "descricao"]

        dados = [
            dict(zip(colunas, row))
            for row in cursor.fetchall()
        ]

    finally:
        liberar(conn)

    return dados


def editar_checklist_pigmentacao(id_registro, nome, data_checklist, status, descricao, armazem_id):

    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        UPDATE checklist_pigmentacao
        SET nome = %s,
            data_checklist = %s,
            status = %s,
            descricao = %s
        WHERE id = %s
        AND armazem_id = %s
        """, (
            nome, data_checklist, status, descricao, id_registro, armazem_id
        ))

        conn.commit()

    finally:
        liberar(conn)

    ler_checklist_pigmentacao.clear()


def excluir_checklist_pigmentacao(id_registro, armazem_id):

    _excluir_checklist(
        "checklist_pigmentacao", id_registro, armazem_id
    )

    ler_checklist_pigmentacao.clear()

# ==================================================
# EQUIPAMENTOS: RESPONSÁVEIS E CARRINHOS FIXOS
# ==================================================

def _ler_responsaveis(tabela, armazem_id):

    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute(f"""
        SELECT id, nome, numero
        FROM {tabela}
        WHERE armazem_id = %s
        ORDER BY nome ASC
        """, (armazem_id,))

        colunas = ["id", "nome", "numero"]

        dados = [
            dict(zip(colunas, row))
            for row in cursor.fetchall()
        ]

    finally:
        liberar(conn)

    return dados


def _adicionar_responsavel(tabela, nome, numero, armazem_id):

    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute(f"""
        INSERT INTO {tabela} (nome, numero, armazem_id)
        VALUES (%s, %s, %s)
        """, (nome, numero, armazem_id))

        conn.commit()

    finally:
        liberar(conn)


def _editar_responsavel(tabela, id_registro, nome, numero, armazem_id):

    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute(f"""
        UPDATE {tabela}
        SET nome = %s,
            numero = %s
        WHERE id = %s
        AND armazem_id = %s
        """, (nome, numero, id_registro, armazem_id))

        conn.commit()

    finally:
        liberar(conn)


def _excluir_responsavel(tabela, id_registro, armazem_id):

    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute(f"""
        DELETE FROM {tabela}
        WHERE id = %s
        AND armazem_id = %s
        """, (id_registro, armazem_id))

        conn.commit()

    finally:
        liberar(conn)


@st.cache_data(ttl=30, show_spinner=False)
def ler_responsaveis_hidraulicos(armazem_id):


    return _ler_responsaveis("responsaveis_hidraulicos", armazem_id)


def adicionar_responsavel_hidraulico(nome, numero, armazem_id):

    _adicionar_responsavel("responsaveis_hidraulicos", nome, numero, armazem_id)
    ler_responsaveis_hidraulicos.clear()


def editar_responsavel_hidraulico(id_registro, nome, numero, armazem_id):

    _editar_responsavel("responsaveis_hidraulicos", id_registro, nome, numero, armazem_id)
    ler_responsaveis_hidraulicos.clear()


def excluir_responsavel_hidraulico(id_registro, armazem_id):

    _excluir_responsavel("responsaveis_hidraulicos", id_registro, armazem_id)
    ler_responsaveis_hidraulicos.clear()


@st.cache_data(ttl=30, show_spinner=False)
def ler_responsaveis_carrinhos(armazem_id):


    return _ler_responsaveis("responsaveis_carrinhos", armazem_id)


def adicionar_responsavel_carrinho(nome, numero, armazem_id):

    _adicionar_responsavel("responsaveis_carrinhos", nome, numero, armazem_id)
    ler_responsaveis_carrinhos.clear()


def editar_responsavel_carrinho(id_registro, nome, numero, armazem_id):

    _editar_responsavel("responsaveis_carrinhos", id_registro, nome, numero, armazem_id)
    ler_responsaveis_carrinhos.clear()


def excluir_responsavel_carrinho(id_registro, armazem_id):

    _excluir_responsavel("responsaveis_carrinhos", id_registro, armazem_id)
    ler_responsaveis_carrinhos.clear()


@st.cache_data(ttl=30, show_spinner=False)
def ler_carrinhos_fixos(armazem_id):


    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        SELECT id, local, numero
        FROM carrinhos_fixos
        WHERE armazem_id = %s
        ORDER BY local ASC, numero ASC
        """, (armazem_id,))

        colunas = ["id", "local", "numero"]

        dados = [
            dict(zip(colunas, row))
            for row in cursor.fetchall()
        ]

    finally:
        liberar(conn)

    return dados


def adicionar_carrinho_fixo(local, numero, armazem_id):

    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        INSERT INTO carrinhos_fixos (local, numero, armazem_id)
        VALUES (%s, %s, %s)
        """, (local, numero, armazem_id))

        conn.commit()

    finally:
        liberar(conn)

    ler_carrinhos_fixos.clear()


def editar_carrinho_fixo(id_registro, local, numero, armazem_id):

    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        UPDATE carrinhos_fixos
        SET local = %s,
            numero = %s
        WHERE id = %s
        AND armazem_id = %s
        """, (local, numero, id_registro, armazem_id))

        conn.commit()

    finally:
        liberar(conn)

    ler_carrinhos_fixos.clear()


def excluir_carrinho_fixo(id_registro, armazem_id):

    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        DELETE FROM carrinhos_fixos
        WHERE id = %s
        AND armazem_id = %s
        """, (id_registro, armazem_id))

        conn.commit()

    finally:
        liberar(conn)

    ler_carrinhos_fixos.clear()

# ==================================================
# USUÁRIOS
# ==================================================

def autenticar(
    usuario,
    senha
):

    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        SELECT
            u.id,
            u.tipo,
            u.trocar_senha,
            u.armazem_id,
            a.nome
        FROM usuarios u
        JOIN armazens a ON a.id = u.armazem_id
        WHERE u.usuario = %s
        AND u.senha = %s
        """, (
            usuario,
            senha
        ))

        resultado = cursor.fetchone()

    finally:
        liberar(conn)

    return resultado


# ==================================================
# SESSÕES ATIVAS (login persistente)
# ==================================================
# Um token é gerado a cada login e colocado na URL da página.
# Como a URL fica salva no navegador, se a página recarregar
# sozinha (F5, uma queda de rede rápida, o navegador reconectando
# o "fio" da aplicação), usamos esse token para reconhecer quem
# era e devolvê-la exatamente para onde estava, sem pedir login
# de novo.

def criar_sessao(usuario, dias_validade=30):

    token = secrets.token_urlsafe(32)

    expira_em = datetime.utcnow() + timedelta(days=dias_validade)

    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        INSERT INTO sessoes_ativas (token, usuario, expira_em)
        VALUES (%s, %s, %s)
        """, (
            token,
            usuario,
            expira_em
        ))

        conn.commit()

    finally:
        liberar(conn)

    return token


def validar_sessao(token):

    if not token:
        return None

    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        SELECT
            u.usuario,
            u.tipo,
            u.trocar_senha,
            u.armazem_id,
            a.nome
        FROM sessoes_ativas s
        JOIN usuarios u ON u.usuario = s.usuario
        JOIN armazens a ON a.id = u.armazem_id
        WHERE s.token = %s
        AND (s.expira_em IS NULL OR s.expira_em > CURRENT_TIMESTAMP)
        """, (token,))

        resultado = cursor.fetchone()

    finally:
        liberar(conn)

    return resultado


def renovar_sessao(token, dias_validade=30):

    if not token:
        return

    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        UPDATE sessoes_ativas
        SET expira_em = %s
        WHERE token = %s
        """, (
            datetime.utcnow() + timedelta(days=dias_validade),
            token
        ))

        conn.commit()

    finally:
        liberar(conn)


def encerrar_sessao(token):

    if not token:
        return

    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        DELETE FROM sessoes_ativas
        WHERE token = %s
        """, (token,))

        conn.commit()

    finally:
        liberar(conn)


def criar_usuario(
    usuario,
    senha,
    armazem_id,
    tipo="usuario"
):

    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        INSERT INTO usuarios
        (
            usuario,
            senha,
            tipo,
            trocar_senha,
            armazem_id
        )
        VALUES (%s, %s, %s, %s, %s)
        """, (
            usuario,
            senha,
            tipo,
            1,
            armazem_id
        ))

        conn.commit()

    finally:
        liberar(conn)

    listar_usuarios.clear()


@st.cache_data(ttl=30, show_spinner=False)
def listar_usuarios(armazem_id):


    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        SELECT
            id,
            usuario,
            tipo,
            ultimo_acesso
        FROM usuarios
        WHERE armazem_id = %s
        ORDER BY usuario
        """, (armazem_id,))

        usuarios = cursor.fetchall()

    finally:
        liberar(conn)

    return usuarios


def atualizar_ultimo_acesso(usuario):

    # Esse "heartbeat" roda sozinho em segundo plano, a cada 120s,
    # para TODA sessão logada (ver render_status_footer no app.py).
    # Por isso precisa ser à prova de falha: se o UPDATE der erro
    # por qualquer instabilidade momentânea do banco, a conexão tem
    # que voltar pro pool mesmo assim (senão, rodando sem parar,
    # acaba vazando as conexões até esgotar o pool) — e o erro não
    # pode nunca derrubar a tela do usuário.

    conn = None

    try:

        conn = conectar()
        cursor = conn.cursor()

        cursor.execute("""
        UPDATE usuarios
        SET ultimo_acesso = CURRENT_TIMESTAMP
        WHERE usuario = %s
        """, (
            usuario,
        ))

        conn.commit()

        listar_usuarios.clear()

    except Exception:

        pass

    finally:

        if conn is not None:
            liberar(conn)


def excluir_usuario(usuario):

    # Nunca permitir apagar o fundador
    if usuario == USUARIO_FUNDADOR:
        return False

    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        DELETE FROM usuarios
        WHERE usuario = %s
        """, (
            usuario,
        ))

        conn.commit()

    finally:
        liberar(conn)

    listar_usuarios.clear()


def excluir_usuario_por_id(uid):

    conn = conectar()

    try:
        cursor = conn.cursor()

        # Nunca permitir apagar o fundador, mesmo por id
        cursor.execute("""
        DELETE FROM usuarios
        WHERE id = %s
        AND usuario IS DISTINCT FROM %s
        """, (
            uid,
            USUARIO_FUNDADOR,
        ))

        conn.commit()

    finally:
        liberar(conn)

    listar_usuarios.clear()

    return True


def alterar_senha(
    usuario,
    nova_senha
):

    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        UPDATE usuarios
        SET
            senha = %s,
            trocar_senha = 0
        WHERE usuario = %s
        """, (
            nova_senha,
            usuario
        ))

        conn.commit()

    finally:
        liberar(conn)


def resetar_senha(
    usuario,
    senha_temporaria
):

    # proteção adicional
    if usuario == USUARIO_FUNDADOR:
        return False

    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        UPDATE usuarios
        SET
            senha = %s,
            trocar_senha = 1
        WHERE usuario = %s
        """, (
            senha_temporaria,
            usuario
        ))

        conn.commit()

    finally:
        liberar(conn)

    return True


# ==================================================
# ARMAZÉNS (MULTI-TENANT)
# ==================================================

@st.cache_data(ttl=30, show_spinner=False)
def listar_armazens():


    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        SELECT id, nome
        FROM armazens
        ORDER BY nome ASC
        """)

        dados = cursor.fetchall()

    finally:
        liberar(conn)

    return dados


def criar_armazem(nome):

    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        INSERT INTO armazens (nome)
        VALUES (%s)
        RETURNING id
        """, (nome,))

        novo_id = cursor.fetchone()[0]

        for nome_rua in RUAS_PADRAO:

            cursor.execute(
                """
                INSERT INTO ruas (armazem_id, nome)
                VALUES (%s, %s)
                ON CONFLICT (armazem_id, nome) DO NOTHING
                """,
                (novo_id, nome_rua)
            )

        conn.commit()

    finally:
        liberar(conn)

    listar_armazens.clear()
    listar_ruas.clear()

    return novo_id


def renomear_armazem(armazem_id, novo_nome):

    conn = conectar()

    try:
        cursor = conn.cursor()

        cursor.execute("""
        UPDATE armazens
        SET nome = %s
        WHERE id = %s
        """, (novo_nome, armazem_id))

        conn.commit()

    finally:
        liberar(conn)

    listar_armazens.clear()
