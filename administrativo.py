import streamlit as st
import banco
import estilos
import remanejamento
import perfil
import pandas as pd
import io
import openpyxl

from datetime import datetime, timezone, time as time_type
from zoneinfo import ZoneInfo


# =====================================================
# PLANILHA DO SAC — leitura e importação de chamados
# =====================================================
# Mapa de coluna (posição, 0-indexada) -> campo da Análise Técnica,
# de acordo com o layout real da planilha "<Mês> <Ano>":
# Data(abertura) | CH | Cliente | NF | Descrição | Cód.Pdt | Produto |
# Tratativa | Data(fechamento) | Hora | Separador | Volume | Carga |
# Região | Motorista | Balança | Conf. | Tipo | Separador(ignorado)

MESES_PT = [
    "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
    "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"
]

VALORES_SEM_NOME = {"", "não", "nao", "sim", "-", "n/a", "na"}


def _texto_celula(valor):

    if valor is None:
        return ""

    return str(valor).strip()


def _data_celula(valor):

    if valor is None or valor == "" or valor == " - ":
        return None

    if isinstance(valor, datetime):
        return valor.date()

    if hasattr(valor, "year") and hasattr(valor, "month") and hasattr(valor, "day"):
        # datetime.date puro (célula formatada só como Data, sem hora) —
        # não tem método .date(), mas já é a data em si.
        return valor

    return None


def _hora_celula(valor):

    if valor is None:
        return None

    if hasattr(valor, "hour"):
        return valor

    return None


def nome_aba_mes_atual():

    agora = estilos.agora_local()

    return f"{MESES_PT[agora.month - 1]} {agora.year}"


def processar_planilha_sac(arquivo_bytes, armazem_id, usuario):
    """
    Lê a cópia da planilha vinculada, acha a aba do mês atual (ex.:
    "Agosto 2026") e registra em Análise Técnica todo chamado (CH)
    que ainda não tinha sido importado, notificando o Separador e o
    Conferente vinculados a cada um. Devolve (importados, nome_aba,
    diagnostico) — ou (None, None, None) se a aba do mês não existir
    na planilha. `diagnostico` é uma lista de strings explicando por
    que cada linha não numerada foi pulada (útil pra debugar planilha
    com formatação diferente do esperado).
    """

    nome_aba = nome_aba_mes_atual()

    workbook = openpyxl.load_workbook(
        io.BytesIO(arquivo_bytes),
        data_only=True
    )

    if nome_aba not in workbook.sheetnames:
        return None, nome_aba, None

    planilha = workbook[nome_aba]

    ja_importados = banco.chamados_ja_importados(armazem_id)

    importados = 0
    diagnostico = []

    for indice_linha, linha in enumerate(
        planilha.iter_rows(min_row=2, values_only=True),
        start=2
    ):

        if linha[0] is None and (len(linha) <= 1 or linha[1] is None):
            # Linha totalmente vazia — não vale nem logar.
            continue

        chamado = _texto_celula(linha[1]) if len(linha) > 1 else ""

        if not chamado:
            diagnostico.append(
                f"Linha {indice_linha}: pulada — coluna CH (B) vazia."
            )
            continue

        if chamado in ja_importados:
            diagnostico.append(
                f"Linha {indice_linha}: pulada — CH '{chamado}' já estava "
                f"importado (duplicado)."
            )
            continue

        data_erro = _data_celula(linha[0])

        if not data_erro:
            diagnostico.append(
                f"Linha {indice_linha}: pulada — CH '{chamado}' com data "
                f"inválida/vazia na coluna A (valor bruto: "
                f"{linha[0]!r}, tipo: {type(linha[0]).__name__})."
            )
            continue

        dados = {
            "tipo_erro": _texto_celula(linha[17]) or "Não informado",
            "data_erro": data_erro,
            "data_fechamento": _data_celula(linha[8]),
            "descricao": _texto_celula(linha[4]),
            "chamado": chamado,
            "cliente": _texto_celula(linha[2]),
            "nota_fiscal": _texto_celula(linha[3]),
            "cod_produto": _texto_celula(linha[5]),
            "produto": _texto_celula(linha[6]),
            "tratativa": _texto_celula(linha[7]),
            "hora": _hora_celula(linha[9]),
            "separador": _texto_celula(linha[10]),
            "volume": _texto_celula(linha[11]),
            "carga": _texto_celula(linha[12]),
            "regiao": _texto_celula(linha[13]),
            "motorista": _texto_celula(linha[14]),
            "balanca": _texto_celula(linha[15]),
            "conferente": _texto_celula(linha[16]),
        }

        vinculos = []

        if dados["separador"]:

            vinculos.append({
                "nome": dados["separador"],
                "papel": "Separador"
            })

        if dados["conferente"].strip().lower() not in VALORES_SEM_NOME:

            vinculos.append({
                "nome": dados["conferente"],
                "papel": "Conferente"
            })

        banco.adicionar_analise_tecnica(
            dados,
            vinculos,
            armazem_id,
            usuario=usuario
        )

        ja_importados.add(chamado)
        importados += 1

        diagnostico.append(
            f"Linha {indice_linha}: importada — CH '{chamado}', "
            f"vínculos: {[v['nome'] + ' (' + v['papel'] + ')' for v in vinculos] or 'nenhum'}."
        )

    return importados, nome_aba, diagnostico


# =====================================================
# CONFIRMAÇÕES (renderizadas dentro de st.popover)
# =====================================================
# Antes usavam @st.dialog, que precisa de um rerun completo do
# app só para abrir o popup (todo o CSS, sidebar e leituras do
# banco rodam de novo antes do aviso aparecer). Um st.popover
# abre na hora, sem ida ao servidor — só o clique de confirmação
# de fato precisa de um rerun.

def perguntar_notificacao(pendente, campo):

    nome_pessoa = pendente[campo]

    rotulo = "Separador(a)" if campo == "separador" else "Conferente"

    st.write(
        f"Deseja notificar o(a) {rotulo} **{nome_pessoa}** sobre o erro?"
    )

    st.caption(
        "Se confirmar, essa ocorrência também vai aparecer no card dessa pessoa."
    )

    c1, c2 = st.columns(2)

    with c1:
        if st.button(
            "✅ Sim",
            width='stretch',
            key=f"notificar_sim_{campo}"
        ):
            st.session_state["pendente_analise_tecnica"][f"notificar_{campo}"] = True
            st.rerun(scope="fragment")

    with c2:
        if st.button(
            "❌ Não",
            width='stretch',
            key=f"notificar_nao_{campo}"
        ):
            st.session_state["pendente_analise_tecnica"][f"notificar_{campo}"] = False
            st.rerun(scope="fragment")


def confirmar_exclusao_rua(nome_rua, armazem_id):

    st.write(
        f"Tem certeza que deseja excluir a **{nome_rua}**?"
    )

    st.caption(
        "A nota, a dupla e o histórico dessa rua no Dashboard também "
        "serão apagados. Essa ação não pode ser desfeita."
    )

    if st.button(
        "✅ Confirmar exclusão",
        width='stretch',
        key=f"confirma_del_rua_{nome_rua}"
    ):
        with estilos.mostrar_processando(f"excluindo '{nome_rua}'..."):
            banco.excluir_rua(nome_rua, armazem_id)
        estilos.notificar_sucesso(f"'{nome_rua}' excluída.")
        st.rerun(scope="fragment")


def confirmar_exclusao_remanejamento(id_item, nome_item, armazem_id):

    st.write(
        f"Tem certeza que deseja excluir **{nome_item}**?"
    )

    st.caption(
        "Essa ação não pode ser desfeita."
    )

    if st.button(
        "✅ Confirmar exclusão",
        width='stretch',
        key=f"confirma_del_remanejamento_{id_item}"
    ):
        with estilos.mostrar_processando(f"excluindo '{nome_item}'..."):
            banco.excluir_remanejamento(id_item, armazem_id)
        estilos.notificar_sucesso(f"'{nome_item}' excluído.")
        st.rerun(scope="fragment")


def confirmar_exclusao_multipla_remanejamento(ids_selecionados, armazem_id):

    st.write(
        f"Tem certeza que deseja excluir **{len(ids_selecionados)}** "
        "prioridade(s) selecionada(s)?"
    )

    st.caption(
        "Essa ação não pode ser desfeita."
    )

    if st.button(
        "✅ Confirmar exclusão",
        width='stretch',
        key="confirma_del_lote_remanejamento"
    ):
        with estilos.mostrar_processando(f"excluindo {len(ids_selecionados)} prioridade(s)..."):
            banco.excluir_remanejamento_lote(ids_selecionados, armazem_id)
        estilos.notificar_sucesso(f"{len(ids_selecionados)} prioridade(s) excluída(s).")
        st.rerun(scope="fragment")


def confirmar_exclusao_remanejamento_agendado(id_item, nome_item, armazem_id):

    st.write(
        f"Tem certeza que deseja excluir o agendamento **{nome_item}**?"
    )

    st.caption(
        "Essa ação não pode ser desfeita."
    )

    if st.button(
        "✅ Confirmar exclusão",
        width='stretch',
        key=f"confirma_del_remanejamento_agendado_{id_item}"
    ):
        with estilos.mostrar_processando(f"excluindo agendamento '{nome_item}'..."):
            banco.excluir_remanejamento_agendado(id_item, armazem_id)
        estilos.notificar_sucesso(f"Agendamento '{nome_item}' excluído.")
        st.rerun(scope="fragment")


def confirmar_exclusao_usuario(uid, nome_usuario):

    st.write(
        f"Tem certeza que deseja excluir o usuário **{nome_usuario}**?"
    )

    st.caption(
        "Essa ação não pode ser desfeita."
    )

    if st.button(
        "✅ Confirmar exclusão",
        width='stretch',
        key=f"confirma_del_usuario_{uid}"
    ):
        with estilos.mostrar_processando(f"excluindo usuário '{nome_usuario}'..."):
            banco.excluir_usuario_por_id(uid)
        estilos.notificar_sucesso(f"usuário '{nome_usuario}' excluído.")
        st.rerun(scope="fragment")


def confirmar_exclusao_usuarios_inconsistentes(ids_selecionados):

    st.write(
        f"Tem certeza que deseja excluir os **{len(ids_selecionados)}** "
        "registro(s) inconsistente(s) (usuários sem nome)?"
    )

    st.caption(
        "Essa ação não pode ser desfeita."
    )

    if st.button(
        "✅ Confirmar exclusão",
        width='stretch',
        key="confirma_del_lote_usuarios_inconsistentes"
    ):
        with estilos.mostrar_processando(f"excluindo {len(ids_selecionados)} registro(s)..."):
            banco.excluir_usuarios_lote(ids_selecionados)
        estilos.notificar_sucesso(f"{len(ids_selecionados)} registro(s) inconsistente(s) excluído(s).")
        st.rerun(scope="fragment")


def confirmar_exclusao_analise_tecnica(id_registro, nome_registro, armazem_id):

    st.write(
        f"Tem certeza que deseja excluir o registro de **{nome_registro}**?"
    )

    st.caption(
        "Essa ação não pode ser desfeita."
    )

    if st.button(
        "✅ Confirmar exclusão",
        width='stretch',
        key=f"confirma_del_analise_{id_registro}"
    ):
        with estilos.mostrar_processando(f"excluindo registro de {nome_registro}..."):
            banco.excluir_analise_tecnica(id_registro, armazem_id)
        estilos.notificar_sucesso(f"registro de {nome_registro} excluído.")
        st.rerun(scope="fragment")


def confirmar_exclusao_multipla_analise_tecnica(ids_selecionados, armazem_id):

    st.write(
        f"Tem certeza que deseja excluir **{len(ids_selecionados)}** "
        "registro(s) selecionado(s)?"
    )

    st.caption(
        "Essa ação não pode ser desfeita."
    )

    if st.button(
        "✅ Confirmar exclusão",
        width='stretch',
        key="confirma_del_lote_analise"
    ):
        with estilos.mostrar_processando(f"excluindo {len(ids_selecionados)} registro(s)..."):
            banco.excluir_analise_tecnica_lote(ids_selecionados, armazem_id)
        estilos.notificar_sucesso(f"{len(ids_selecionados)} registro(s) excluído(s).")
        st.rerun(scope="fragment")


def confirmar_exclusao_auditoria(id_registro, nome_registro, armazem_id):

    st.write(
        f"Tem certeza que deseja excluir o registro de auditoria de **{nome_registro}**?"
    )

    st.caption(
        "Essa ação não pode ser desfeita."
    )

    if st.button(
        "✅ Confirmar exclusão",
        width='stretch',
        key=f"confirma_del_auditoria_{id_registro}"
    ):
        with estilos.mostrar_processando(f"excluindo registro de {nome_registro}..."):
            banco.excluir_auditoria(id_registro, armazem_id)
        estilos.notificar_sucesso(f"registro de {nome_registro} excluído.")
        st.rerun(scope="fragment")


def confirmar_exclusao_multipla_auditoria(ids_selecionados, armazem_id):

    st.write(
        f"Tem certeza que deseja excluir **{len(ids_selecionados)}** "
        "registro(s) de auditoria selecionado(s)?"
    )

    st.caption(
        "Essa ação não pode ser desfeita."
    )

    if st.button(
        "✅ Confirmar exclusão",
        width='stretch',
        key="confirma_del_lote_auditoria"
    ):
        with estilos.mostrar_processando(f"excluindo {len(ids_selecionados)} registro(s)..."):
            banco.excluir_auditoria_lote(ids_selecionados, armazem_id)
        estilos.notificar_sucesso(f"{len(ids_selecionados)} registro(s) excluído(s).")
        st.rerun(scope="fragment")


def confirmar_exclusao_pessoa_rotativo(id_pessoa, nome_pessoa, armazem_id):

    st.write(
        f"Tem certeza que deseja remover **{nome_pessoa}** do rodízio?"
    )

    st.caption(
        "Essa ação não pode ser desfeita."
    )

    if st.button(
        "✅ Confirmar exclusão",
        width='stretch',
        key=f"confirma_del_pessoa_rot_{id_pessoa}"
    ):
        with estilos.mostrar_processando(f"removendo '{nome_pessoa}'..."):
            banco.excluir_pessoa_rotativo(id_pessoa, armazem_id)
        estilos.notificar_sucesso(f"'{nome_pessoa}' removido(a) do rodízio.")
        st.rerun(scope="fragment")


def confirmar_exclusao_atividade_rotativo(id_atividade, nome_atividade, armazem_id):

    st.write(
        f"Tem certeza que deseja remover a atividade **{nome_atividade}**?"
    )

    st.caption(
        "Essa ação não pode ser desfeita."
    )

    if st.button(
        "✅ Confirmar exclusão",
        width='stretch',
        key=f"confirma_del_ativ_rot_{id_atividade}"
    ):
        with estilos.mostrar_processando(f"removendo '{nome_atividade}'..."):
            banco.excluir_atividade_rotativo(id_atividade, armazem_id)
        estilos.notificar_sucesso(f"'{nome_atividade}' removida do rodízio.")
        st.rerun(scope="fragment")


def confirmar_exclusao_responsavel_hidraulico(id_registro, nome, numero, armazem_id):

    st.write(
        f"Tem certeza que deseja remover **{nome}** como responsável pelo "
        f"Hidráulico {numero}?"
    )

    st.caption(
        "Essa ação não pode ser desfeita."
    )

    if st.button(
        "✅ Confirmar exclusão",
        width='stretch',
        key=f"confirma_del_resp_hid_{id_registro}"
    ):
        with estilos.mostrar_processando(f"removendo '{nome}'..."):
            banco.excluir_responsavel_hidraulico(id_registro, armazem_id)
        estilos.notificar_sucesso(f"'{nome}' removido(a).")
        st.rerun(scope="fragment")


def confirmar_exclusao_responsavel_carrinho(id_registro, nome, numero, armazem_id):

    st.write(
        f"Tem certeza que deseja remover **{nome}** como responsável pelo "
        f"Carrinho {numero}?"
    )

    st.caption(
        "Essa ação não pode ser desfeita."
    )

    if st.button(
        "✅ Confirmar exclusão",
        width='stretch',
        key=f"confirma_del_resp_car_{id_registro}"
    ):
        with estilos.mostrar_processando(f"removendo '{nome}'..."):
            banco.excluir_responsavel_carrinho(id_registro, armazem_id)
        estilos.notificar_sucesso(f"'{nome}' removido(a).")
        st.rerun(scope="fragment")


def confirmar_exclusao_carrinho_fixo(id_registro, local, numero, armazem_id):

    st.write(
        f"Tem certeza que deseja remover o Carrinho **{numero}** "
        f"do local **{local}**?"
    )

    st.caption(
        "Essa ação não pode ser desfeita."
    )

    if st.button(
        "✅ Confirmar exclusão",
        width='stretch',
        key=f"confirma_del_carrinho_fixo_{id_registro}"
    ):
        with estilos.mostrar_processando(f"removendo carrinho {numero}..."):
            banco.excluir_carrinho_fixo(id_registro, armazem_id)
        estilos.notificar_sucesso(f"Carrinho {numero} removido de {local}.")
        st.rerun(scope="fragment")


def render():

    estilos.exibir_notificacao_pendente()

    banco.inicializar_banco()

    usuario_logado = st.session_state.get(
        "usuario",
        ""
    )

    fundador = usuario_logado.startswith(
        "Fundador."
    )

    gestao = usuario_logado.startswith(
        "Gestao."
    )

    admin_master = fundador or gestao

    armazem_id_atual = st.session_state.get(
        "armazem_visualizado_id",
        st.session_state.get("armazem_id")
    )

    estilos.cabecalho_pagina(
        "⚙️",
        "Painel Administrativo Luxiz IA",
        "Gerenciamento central da operação",
        cor="#64748b"
    )

    st.divider()

    total_ruas = len(
        banco.listar_ruas(armazem_id_atual)
    )

    total_remanejamentos = len(
        banco.ler_remanejamentos(armazem_id_atual)
    )

    historico = banco.ler_historico_sac(armazem_id_atual)

    ultimo_sac = 0

    if historico:
        ultimo_sac = historico[-1][1]

    c1, c2, c3 = st.columns(3)

    with c1:
        st.metric(
            "📍 Ruas",
            total_ruas
        )

    with c2:
        st.metric(
            "⚡ Remanejamentos",
            total_remanejamentos
        )

    with c3:
        st.metric(
            "📢 Reclamações",
            ultimo_sac
        )

    st.divider()

    abas = [
        "📊 Dashboard",
        "⚡ Remanejamento",
        "😊 SAC",
        "🎯 Auditoria",
        "🔄 Rodízio",
        "🧰 Equipamentos"
    ]

    if admin_master:
        abas.append("👥 Usuários")

    if fundador:
        abas.append("🏢 Armazéns")

    tabs = st.tabs(abas)

    tab_dashboard = tabs[0]
    tab_remanejamento = tabs[1]
    tab_sac = tabs[2]
    tab_auditoria = tabs[3]
    tab_rotativo = tabs[4]
    tab_equipamentos = tabs[5]

    if admin_master:
        tab_usuarios = tabs[6]

    if fundador:
        tab_armazens = tabs[7]

    # =====================================================
    # DASHBOARD
    # =====================================================

    with tab_dashboard:

        @st.fragment
        def _fragmento_dashboard():

            estilos.exibir_notificacao_pendente()


            st.subheader(
                "Atualização das Equipes"
            )

            ruas = banco.listar_ruas(armazem_id_atual)

            if not ruas:

                st.warning(
                    "Nenhuma rua cadastrada ainda. Crie uma rua no "
                    "campo '➕ Criar Rua' logo abaixo."
                )

            else:

                dados = banco.ler_tudo(armazem_id_atual)

                rua = st.selectbox(
                    "Selecione a Rua",
                    ruas
                )

                info = dados.get(
                    rua,
                    {
                        "nota": 0,
                        "dupla": ""
                    }
                )

                dupla = st.text_input(
                    "Nome da dupla",
                    value=info["dupla"]
                )

                nota = st.slider(
                    "Nota",
                    0.0,
                    5.0,
                    float(info["nota"]),
                    0.1
                )

                if st.button(
                    "💾 Salvar Dashboard"
                ):

                    with estilos.mostrar_processando(f"Dashboard da {rua}..."):
                        banco.salvar_dados(
                            rua,
                            nota,
                            dupla,
                            armazem_id_atual,
                            usuario=usuario_logado
                        )

                    estilos.notificar_sucesso(f"Dashboard da {rua} atualizado.")
                    st.rerun(scope="fragment")

            st.divider()

            # =====================================================
            # CRIAR / EXCLUIR RUA (guardado em campos à parte para
            # não bagunçar a atualização das equipes acima)
            # =====================================================

            with st.expander(
                "➕ Criar Rua"
            ):

                nova_rua = st.text_input(
                    "Nome da nova rua",
                    key="input_nova_rua",
                    placeholder="Ex.: Rua 08"
                )

                if st.button(
                    "➕ Criar Rua",
                    key="botao_criar_rua"
                ):

                    nome_novo = nova_rua.strip()

                    if not nome_novo:

                        st.warning(
                            "Digite um nome para a rua."
                        )

                    elif nome_novo.title() in [r.title() for r in ruas]:

                        st.warning(
                            f"Já existe uma rua chamada '{nome_novo}'."
                        )

                    else:

                        with estilos.mostrar_processando(f"criando '{nome_novo}'..."):
                            banco.criar_rua(
                                nome_novo,
                                armazem_id_atual
                            )

                        estilos.notificar_sucesso(f"Rua '{nome_novo}' criada.")
                        st.rerun(scope="fragment")

            with st.expander(
                "🗑️ Excluir Rua"
            ):

                if not ruas:

                    st.caption(
                        "Nenhuma rua cadastrada ainda."
                    )

                else:

                    rua_para_excluir = st.selectbox(
                        "Selecione a rua a excluir",
                        ruas,
                        key="select_excluir_rua"
                    )

                    st.caption(
                        "Isso também apaga a nota, a dupla e o histórico "
                        "dessa rua no Dashboard. Essa ação não pode ser "
                        "desfeita."
                    )

                    with st.popover(
                        f"🗑️ Excluir '{rua_para_excluir}'"
                    ):

                        confirmar_exclusao_rua(
                            rua_para_excluir,
                            armazem_id_atual
                        )

        _fragmento_dashboard()

    # =====================================================
    # REMANEJAMENTO
    # =====================================================

    with tab_remanejamento:

        @st.fragment
        def _fragmento_remanejamento():

            estilos.exibir_notificacao_pendente()


            st.subheader(
                "Gerenciar Prioridades"
            )

            col1, col2 = st.columns([3, 2])

            with col1:
                novo_item = st.text_input(
                    "Nova prioridade"
                )

            with col2:
                prioridade = st.selectbox(
                    "Prioridade",
                    [
                        "Normal",
                        "Média",
                        "Alta"
                    ]
                )

            if st.button(
                "➕ Adicionar Prioridade"
            ):

                if novo_item:

                    with estilos.mostrar_processando(f"adicionando '{novo_item}'..."):
                        banco.adicionar_remanejamento(
                            novo_item,
                            armazem_id_atual,
                            prioridade,
                            usuario=usuario_logado
                        )

                    estilos.notificar_sucesso(f"'{novo_item}' adicionado.")
                    st.rerun(scope="fragment")

            st.divider()

            itens = banco.ler_remanejamentos(armazem_id_atual)

            if not itens:
                st.info(
                    "Nenhuma prioridade cadastrada."
                )

            ids_selecionados_remanejamento = []

            for item in itens:

                c0, c1, c2 = st.columns([0.6, 7.4, 1])

                with c0:

                    marcado_remanejamento = st.checkbox(
                        "selecionar",
                        key=f"select_remanejamento_{item['id']}",
                        label_visibility="collapsed"
                    )

                    if marcado_remanejamento:
                        ids_selecionados_remanejamento.append(item["id"])

                with c1:

                    if item["prioridade"] == "Alta":

                        st.error(
                            f"🚨 {item['nome']} • PRIORIDADE ALTA"
                        )

                    elif item["prioridade"] == "Média":

                        st.warning(
                            f"⚠️ {item['nome']} • PRIORIDADE MÉDIA"
                        )

                    else:

                        st.success(
                            f"✅ {item['nome']} • PRIORIDADE NORMAL"
                        )

                with c2:

                    with st.popover("❌", key=f"pop_del_reman_{item['id']}"):

                        confirmar_exclusao_remanejamento(
                            item["id"],
                            item["nome"],
                            armazem_id_atual
                        )

            if ids_selecionados_remanejamento:

                st.write("")

                with st.popover(
                    f"🗑️ Excluir {len(ids_selecionados_remanejamento)} selecionado(s)"
                ):

                    confirmar_exclusao_multipla_remanejamento(
                        ids_selecionados_remanejamento,
                        armazem_id_atual
                    )

        _fragmento_remanejamento()

        st.divider()

        @st.fragment
        def _fragmento_remanejamento_agendado():

            estilos.exibir_notificacao_pendente()

            st.subheader(
                "⏰ Agendamento por Horário"
            )

            st.caption(
                "Cadastre docas/prioridades que devem entrar e sair do "
                "painel sozinhas, de acordo com o horário e os dias da "
                "semana. Elas aparecem juntas com as prioridades manuais "
                "enquanto estiverem dentro da janela configurada."
            )

            col1, col2 = st.columns([3, 2])

            with col1:
                novo_item_agendado = st.text_input(
                    "Nome (ex.: Doca 11)",
                    key="reman_agendado_nome"
                )

            with col2:
                prioridade_agendada = st.selectbox(
                    "Prioridade",
                    [
                        "Normal",
                        "Média",
                        "Alta"
                    ],
                    key="reman_agendado_prioridade"
                )

            col3, col4 = st.columns(2)

            with col3:
                hora_inicio_agendada = st.time_input(
                    "Horário de entrada",
                    value=time_type(8, 0),
                    key="reman_agendado_hora_inicio"
                )

            with col4:
                hora_fim_agendada = st.time_input(
                    "Horário de saída",
                    value=time_type(13, 0),
                    key="reman_agendado_hora_fim"
                )

            dias_disponiveis = [
                "Segunda", "Terça", "Quarta", "Quinta",
                "Sexta", "Sábado", "Domingo"
            ]

            dias_selecionados = st.multiselect(
                "Dias da semana",
                dias_disponiveis,
                default=["Segunda", "Terça", "Quarta", "Quinta", "Sexta"],
                key="reman_agendado_dias"
            )

            if st.button(
                "➕ Adicionar Agendamento",
                key="reman_agendado_botao_add"
            ):

                if not novo_item_agendado.strip():

                    st.warning(
                        "Informe o nome do item (ex.: Doca 11)."
                    )

                elif not dias_selecionados:

                    st.warning(
                        "Selecione ao menos um dia da semana."
                    )

                elif hora_inicio_agendada == hora_fim_agendada:

                    st.warning(
                        "O horário de entrada e saída não podem ser iguais."
                    )

                else:

                    dias_semana_indices = [
                        dias_disponiveis.index(dia)
                        for dia in dias_selecionados
                    ]

                    with estilos.mostrar_processando(
                        f"agendando '{novo_item_agendado.strip()}'..."
                    ):
                        banco.criar_remanejamento_agendado(
                            novo_item_agendado.strip(),
                            prioridade_agendada,
                            hora_inicio_agendada,
                            hora_fim_agendada,
                            dias_semana_indices,
                            armazem_id_atual,
                            usuario=usuario_logado
                        )

                    estilos.notificar_sucesso(
                        f"'{novo_item_agendado.strip()}' agendado."
                    )
                    st.rerun(scope="fragment")

            st.divider()

            agendados = banco.ler_remanejamentos_agendados(armazem_id_atual)

            if not agendados:

                st.info(
                    "Nenhum agendamento por horário cadastrado."
                )

            else:

                agora = estilos.agora_local()

                for agendamento in agendados:

                    ativo_agora = remanejamento.agendamento_esta_ativo(
                        agendamento,
                        agora
                    )

                    c1, c2 = st.columns([7.4, 1])

                    with c1:

                        with st.container(border=True):

                            st.markdown(
                                f"**{agendamento['nome']}** • "
                                f"{agendamento['prioridade']}"
                            )

                            st.caption(
                                remanejamento.descricao_janela(agendamento)
                            )

                            if ativo_agora:

                                st.success(
                                    "🟢 Ativo agora — visível no painel."
                                )

                            else:

                                st.caption(
                                    "⚪ Fora do horário no momento."
                                )

                            if agendamento.get("criado_por"):

                                st.caption(
                                    f"👤 Criado por {agendamento['criado_por']}"
                                )

                    with c2:

                        with st.popover("❌", key=f"pop_del_reman_ag_{agendamento['id']}"):

                            confirmar_exclusao_remanejamento_agendado(
                                agendamento["id"],
                                agendamento["nome"],
                                armazem_id_atual
                            )

        _fragmento_remanejamento_agendado()

    # =====================================================
    # SAC
    # =====================================================

    with tab_sac:

        @st.fragment
        def _fragmento_sac():

            estilos.exibir_notificacao_pendente()


            st.subheader(
                "Atualização Mensal do SAC"
            )

            with st.form(
                "form_sac"
            ):

                reclamacoes = st.number_input(
                    "Reclamações",
                    min_value=0,
                    step=1
                )

                meta = st.number_input(
                    "Meta",
                    min_value=0,
                    step=1
                )

                salvar = st.form_submit_button(
                    "💾 Salvar SAC"
                )

                if salvar:

                    with estilos.mostrar_processando("dados do SAC..."):
                        banco.atualizar_sac_mensal(
                            reclamacoes,
                            meta,
                            armazem_id_atual,
                            usuario=usuario_logado
                        )

                    estilos.notificar_sucesso("SAC atualizado.")
                    st.rerun(scope="fragment")

            st.divider()

            # -----------------------------------------------
            # VINCULAR PLANILHA
            # -----------------------------------------------

            st.subheader("🔗 Vincular planilha")

            st.caption(
                "Envie o Excel de chamados. O app procura, dentro dele, "
                f"a aba do mês atual (**{nome_aba_mes_atual()}**) e "
                "registra em Análise Técnica todo chamado (CH) que "
                "ainda não existir no sistema, notificando quem estiver "
                "vinculado (Separador/Conferente)."
            )

            planilha_vinculada = banco.ler_planilha_sac(armazem_id_atual)

            if planilha_vinculada:

                atualizado_em_local = estilos.horario_local(
                    planilha_vinculada["enviado_em"]
                )

                st.caption(
                    f"📎 Planilha atual: **{planilha_vinculada['nome_arquivo']}** "
                    f"— enviada por {planilha_vinculada['enviado_por'] or 'desconhecido'} "
                    f"em {atualizado_em_local.strftime('%d/%m/%Y %H:%M')}"
                )

            novo_arquivo_sac = st.file_uploader(
                "Planilha de chamados (.xlsx)",
                type=["xlsx"],
                key="upload_planilha_sac"
            )

            col_vincular, col_processar = st.columns(2)

            with col_vincular:

                if st.button(
                    "📎 Vincular planilha",
                    width='stretch',
                    disabled=novo_arquivo_sac is None
                ):

                    with estilos.mostrar_processando("vinculando planilha..."):

                        banco.salvar_planilha_sac(
                            armazem_id_atual,
                            novo_arquivo_sac.name,
                            novo_arquivo_sac.getvalue(),
                            usuario_logado
                        )

                        importados, nome_aba, diagnostico = processar_planilha_sac(
                            novo_arquivo_sac.getvalue(),
                            armazem_id_atual,
                            usuario_logado
                        )

                    st.session_state["diagnostico_planilha_sac"] = diagnostico

                    if importados is None:

                        st.warning(
                            f"Planilha vinculada, mas não encontrei a aba "
                            f"'{nome_aba}' nela — confira o nome da aba do "
                            f"mês atual."
                        )

                    else:

                        estilos.notificar_sucesso(
                            f"Planilha vinculada. {importados} chamado(s) "
                            f"novo(s) importado(s) de '{nome_aba}'."
                        )

                    st.rerun(scope="fragment")

            with col_processar:

                if st.button(
                    "🔄 Reprocessar planilha atual",
                    width='stretch',
                    disabled=not planilha_vinculada
                ):

                    with estilos.mostrar_processando("procurando chamados novos..."):
                        importados, nome_aba, diagnostico = processar_planilha_sac(
                            planilha_vinculada["conteudo"],
                            armazem_id_atual,
                            usuario_logado
                        )

                    st.session_state["diagnostico_planilha_sac"] = diagnostico

                    if importados is None:

                        st.warning(
                            f"Não encontrei a aba '{nome_aba}' na planilha "
                            f"vinculada."
                        )

                    elif importados == 0:

                        st.info("Nenhum chamado novo encontrado.")

                    else:

                        estilos.notificar_sucesso(
                            f"{importados} chamado(s) novo(s) importado(s) "
                            f"de '{nome_aba}'."
                        )

                    st.rerun(scope="fragment")

            diagnostico_salvo = st.session_state.get("diagnostico_planilha_sac")

            if diagnostico_salvo:

                with st.expander(
                    f"🔬 Diagnóstico do último processamento "
                    f"({len(diagnostico_salvo)} linha(s) lida(s))"
                ):

                    for linha_diagnostico in diagnostico_salvo:
                        st.caption(linha_diagnostico)

            st.caption(
                "⚠️ O app não enxerga o arquivo no seu computador — ele "
                "trabalha sempre em cima da última cópia enviada aqui. "
                "Editou a planilha local? Envie ela de novo em "
                "'Vincular planilha' pra atualizar."
            )

            st.divider()

            st.subheader(
                "🔍 Análise Técnica"
            )

            st.caption(
                "Registre os detalhes completos da ocorrência."
            )

            TIPOS_ERRO_SAC = [
                "Pigmentação",
                "Componente",
                "Contagem",
                "Deixou no Picking",
                "Impróprio",
                "Inversão de Doca",
                "Inversão de Etiqueta",
                "Inversão de Picking",
                "Inversão de Produto"
            ]

            TRATATIVA_OPCOES = [
                "Crédito disponível",
                "Minuta",
                "Coleta",
                "Pedido"
            ]

            BALANCA_OPCOES = [
                "Sim",
                "Não"
            ]

            # -----------------------------------------------
            # Fluxo de confirmação de notificação
            # (roda antes do formulário para não perder o
            # estado pendente entre reruns)
            # -----------------------------------------------

            pendente = st.session_state.get(
                "pendente_analise_tecnica"
            )

            if pendente:

                if pendente.get("separador") and pendente.get("notificar_separador") is None:

                    with st.container(border=True):

                        st.markdown("**🔔 Confirmar notificação**")

                        perguntar_notificacao(
                            pendente,
                            campo="separador"
                        )

                elif pendente.get("conferente") and pendente.get("notificar_conferente") is None:

                    with st.container(border=True):

                        st.markdown("**🔔 Confirmar notificação**")

                        perguntar_notificacao(
                            pendente,
                            campo="conferente"
                        )

                else:

                    vinculos = []

                    if pendente.get("separador") and pendente.get("notificar_separador"):

                        vinculos.append({
                            "nome": pendente["separador"],
                            "papel": "Separador"
                        })

                    if pendente.get("conferente") and pendente.get("notificar_conferente"):

                        vinculos.append({
                            "nome": pendente["conferente"],
                            "papel": "Conferente"
                        })

                    # Deduplica por nome (ignorando maiúsculas/minúsculas):
                    # se a mesma pessoa aparecer em mais de um campo,
                    # essa ocorrência conta 1 vez só, com os papéis mesclados.
                    vinculos_por_nome = {}

                    for vinculo in vinculos:

                        chave = vinculo["nome"].strip().title()

                        if chave in vinculos_por_nome:

                            if vinculo["papel"] not in vinculos_por_nome[chave]["papel"]:

                                vinculos_por_nome[chave]["papel"] += f" e {vinculo['papel']}"

                        else:

                            vinculos_por_nome[chave] = {
                                "nome": vinculo["nome"],
                                "papel": vinculo["papel"]
                            }

                    vinculos = list(vinculos_por_nome.values())

                    rotulo_ocorrencia = (
                        pendente.get("separador")
                        or pendente.get("conferente")
                        or pendente.get("chamado")
                        or "nova ocorrência"
                    )

                    with estilos.mostrar_processando(f"registrando {rotulo_ocorrencia}..."):
                        banco.adicionar_analise_tecnica(
                            pendente,
                            vinculos,
                            armazem_id_atual,
                            usuario=usuario_logado
                        )

                    del st.session_state["pendente_analise_tecnica"]

                    st.session_state["chk_chamado_fechado_novo"] = False

                    estilos.notificar_sucesso(f"{rotulo_ocorrencia} registrada.")
                    st.rerun(scope="fragment")

            # O checkbox de "chamado já fechado" fica fora do st.form:
            # widgets dentro de um form só atualizam a tela no submit,
            # então não daria pra revelar o campo de data de fechamento
            # na hora. Aqui fora, ele reage na hora (rerun do fragmento).

            chamado_ja_fechado = st.checkbox(
                "Chamado já foi fechado? (informe a data de fechamento abaixo)",
                key="chk_chamado_fechado_novo"
            )

            with st.form(
                "form_analise_tecnica"
            ):

                col2, col3, col3b = st.columns(3)

                with col2:
                    tipo_erro = st.selectbox(
                        "Tipo",
                        TIPOS_ERRO_SAC
                    )

                with col3:
                    data_abertura = st.date_input(
                        "Data de abertura do chamado"
                    )

                with col3b:

                    if chamado_ja_fechado:

                        data_fechamento = st.date_input(
                            "Data de fechamento do chamado",
                            key="data_fechamento_novo_chamado"
                        )

                    else:

                        data_fechamento = None

                        st.caption(
                            "🟡 Sem data de fechamento, o chamado ficará "
                            "**Pendente**. A pessoa vinculada já pode ser "
                            "notificada normalmente — a data de fechamento "
                            "pode ser informada depois, na listagem abaixo."
                        )

                col4, col5, col6 = st.columns(3)

                with col4:
                    chamado = st.text_input(
                        "Chamado"
                    )

                with col5:
                    cliente = st.text_input(
                        "Cliente"
                    )

                with col6:
                    nota_fiscal = st.text_input(
                        "Nota Fiscal"
                    )

                col7, col8, col9 = st.columns(3)

                with col7:
                    cod_produto = st.text_input(
                        "Cód Produto"
                    )

                with col8:
                    produto = st.text_input(
                        "Produto"
                    )

                with col9:
                    hora = st.time_input(
                        "Hora"
                    )

                col10, col11, col12 = st.columns(3)

                with col10:
                    separador = st.text_input(
                        "Separador"
                    )

                with col11:
                    conferente = st.text_input(
                        "Conferente"
                    )

                with col12:
                    balanca = st.selectbox(
                        "Balança",
                        BALANCA_OPCOES
                    )

                col13, col14, col15 = st.columns(3)

                with col13:
                    volume = st.text_input(
                        "Volume"
                    )

                with col14:
                    carga = st.text_input(
                        "Carga"
                    )

                with col15:
                    regiao = st.text_input(
                        "Região"
                    )

                col16, col17 = st.columns(2)

                with col16:
                    motorista = st.text_input(
                        "Motorista"
                    )

                with col17:
                    tratativa = st.selectbox(
                        "Tratativa",
                        TRATATIVA_OPCOES
                    )

                descricao_erro = st.text_area(
                    "Descrição do ocorrido"
                )

                registrar = st.form_submit_button(
                    "➕ Registrar Análise Técnica"
                )

                if registrar:

                    st.session_state["pendente_analise_tecnica"] = {
                        "tipo_erro": tipo_erro,
                        "data_erro": data_abertura,
                        "data_fechamento": data_fechamento if chamado_ja_fechado else None,
                        "descricao": descricao_erro,
                        "chamado": chamado,
                        "cliente": cliente,
                        "nota_fiscal": nota_fiscal,
                        "cod_produto": cod_produto,
                        "produto": produto,
                        "tratativa": tratativa,
                        "hora": hora,
                        "separador": separador if separador else None,
                        "volume": volume,
                        "carga": carga,
                        "regiao": regiao,
                        "motorista": motorista,
                        "balanca": balanca,
                        "conferente": conferente if conferente else None,
                        "notificar_separador": None if separador else False,
                        "notificar_conferente": None if conferente else False
                    }

                    st.rerun(scope="fragment")

            st.divider()

            registros_tecnica = banco.ler_analise_tecnica(armazem_id_atual)

            if registros_tecnica:

                CAMPOS_BUSCA_ANALISE = {
                    "Tipo": "tipo_erro",
                    "Chamado": "chamado",
                    "Cliente": "cliente",
                    "Nota Fiscal": "nota_fiscal",
                    "Cód Produto": "cod_produto",
                    "Produto": "produto",
                    "Separador": "separador",
                    "Conferente": "conferente",
                    "Motorista": "motorista",
                    "Volume": "volume",
                    "Carga": "carga",
                    "Região": "regiao",
                    "Tratativa": "tratativa",
                    "Balança": "balanca",
                    "Data de abertura": "data_erro",
                    "Data de fechamento": "data_fechamento",
                }

                col_busca1, col_busca2 = st.columns([1, 2])

                with col_busca1:

                    campo_busca_rotulo = st.selectbox(
                        "🔍 Procurar por:",
                        list(CAMPOS_BUSCA_ANALISE.keys()),
                        key="campo_busca_analise_tecnica"
                    )

                with col_busca2:

                    texto_busca = st.text_input(
                        f"Digite para filtrar por {campo_busca_rotulo.lower()}",
                        key="texto_busca_analise_tecnica",
                        placeholder="Digite aqui para filtrar... (ex: 15/07 para datas)"
                    )

                campo_busca_coluna = CAMPOS_BUSCA_ANALISE[campo_busca_rotulo]

                def _valor_busca_como_texto(registro, coluna):

                    valor = registro.get(coluna)

                    if valor is None:
                        return ""

                    if hasattr(valor, "strftime"):
                        return valor.strftime("%d/%m/%Y")

                    return str(valor)

                if texto_busca.strip():

                    termo_busca = texto_busca.strip().lower()

                    registros_tecnica = [
                        registro for registro in registros_tecnica
                        if termo_busca in _valor_busca_como_texto(
                            registro,
                            campo_busca_coluna
                        ).lower()
                    ]

                    if not registros_tecnica:

                        st.warning(
                            f"Nenhum chamado encontrado para \"{texto_busca}\" "
                            f"em {campo_busca_rotulo}."
                        )

                st.divider()

                MESES_COMPLETOS = [
                    "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
                    "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"
                ]

                por_mes = {}

                for registro in registros_tecnica:

                    chave_mes = registro["data_erro"].strftime("%Y-%m")

                    por_mes.setdefault(
                        chave_mes,
                        []
                    ).append(registro)

                meses_ordenados = sorted(
                    por_mes.keys(),
                    reverse=True
                )

                st.caption(
                    "Marque as caixinhas para excluir vários de uma vez, ou "
                    "clique em ❌ para excluir um lançamento só. Chamados sem "
                    "data de fechamento aparecem como 🟡 Pendente."
                )

                for indice_mes, chave_mes in enumerate(meses_ordenados):

                    registros_mes = por_mes[chave_mes]

                    ano_mes, numero_mes = chave_mes.split("-")

                    nome_mes = MESES_COMPLETOS[int(numero_mes) - 1]

                    pendentes_mes = len(
                        [
                            r for r in registros_mes
                            if not r.get("data_fechamento")
                        ]
                    )

                    rotulo_expander = (
                        f"🗓️ Chamados de {nome_mes}/{ano_mes} "
                        f"({len(registros_mes)})"
                    )

                    if pendentes_mes:

                        rotulo_expander += f" • 🟡 {pendentes_mes} pendente(s)"

                    with st.expander(
                        rotulo_expander,
                        expanded=(indice_mes == 0)
                    ):

                        ids_selecionados_mes = []

                        for registro in registros_mes:

                            c0, c1, c2 = st.columns([0.6, 7.4, 1])

                            with c0:

                                marcado = st.checkbox(
                                    "selecionar",
                                    key=f"select_analise_{registro['id']}",
                                    label_visibility="collapsed"
                                )

                                if marcado:
                                    ids_selecionados_mes.append(registro["id"])

                            with c1:

                                identificador = (
                                    registro.get("separador")
                                    or registro.get("conferente")
                                    or registro.get("chamado")
                                    or "Ocorrência"
                                )

                                chamado_registro = registro.get("chamado")
                                fechamento_registro = registro.get("data_fechamento")

                                status_rotulo = (
                                    "🟢 Finalizado"
                                    if fechamento_registro
                                    else "🟡 Pendente"
                                )

                                st.caption(
                                    f"👤 {identificador} • {registro['tipo_erro']} • "
                                    f"Abertura {registro['data_erro'].strftime('%d/%m/%Y')}"
                                    + (
                                        f" • Fechamento {fechamento_registro.strftime('%d/%m/%Y')}"
                                        if fechamento_registro
                                        else ""
                                    )
                                    + (f" • Chamado {chamado_registro}" if chamado_registro else "")
                                    + f" • {status_rotulo}"
                                )

                                if not fechamento_registro:

                                    with st.popover(
                                        "🔒 Informar data de fechamento",
                                        key=f"pop_fechar_analise_{registro['id']}"
                                    ):

                                        nova_data_fechamento = st.date_input(
                                            "Data de fechamento do chamado",
                                            key=f"input_fechamento_{registro['id']}"
                                        )

                                        if st.button(
                                            "✅ Finalizar chamado",
                                            key=f"confirma_fechar_{registro['id']}"
                                        ):

                                            with estilos.mostrar_processando("finalizando chamado..."):
                                                banco.finalizar_analise_tecnica(
                                                    registro["id"],
                                                    nova_data_fechamento,
                                                    armazem_id_atual
                                                )

                                            estilos.notificar_sucesso("Chamado finalizado.")
                                            st.rerun(scope="fragment")

                            with c2:

                                with st.popover("❌", key=f"pop_del_analise_{registro['id']}"):

                                    confirmar_exclusao_analise_tecnica(
                                        registro["id"],
                                        identificador,
                                        armazem_id_atual
                                    )

                        if ids_selecionados_mes:

                            st.write("")

                            with st.popover(
                                f"🗑️ Excluir {len(ids_selecionados_mes)} selecionado(s)",
                                key=f"pop_del_lote_{chave_mes}"
                            ):

                                confirmar_exclusao_multipla_analise_tecnica(
                                    ids_selecionados_mes,
                                    armazem_id_atual
                                )

                        st.divider()

                        # A montagem do Excel só roda quando a pessoa clica em
                        # "Gerar" (e fica guardada no session_state depois
                        # disso). Gerar isso pra todos os meses de uma vez,
                        # toda vez que a página carrega, é o que estava
                        # deixando a entrada no Administrativo lenta.

                        chave_excel_pronto = f"excel_pronto_{chave_mes}"

                        if st.button(
                            f"📊 Gerar Excel de {nome_mes}/{ano_mes}",
                            key=f"gerar_excel_{chave_mes}"
                        ):

                            df_exportar_mes = pd.DataFrame(registros_mes)

                            if "vinculos_notificados" in df_exportar_mes.columns:

                                df_exportar_mes["vinculos_notificados"] = df_exportar_mes["vinculos_notificados"].apply(
                                    lambda lista: ", ".join(
                                        f"{item['nome']} ({item['papel']})" for item in lista
                                    ) if lista else ""
                                )

                            if "data_fechamento" in df_exportar_mes.columns:

                                df_exportar_mes["status"] = df_exportar_mes["data_fechamento"].apply(
                                    lambda data: "Finalizado" if data else "Pendente"
                                )

                            buffer_excel_mes = io.BytesIO()

                            with pd.ExcelWriter(buffer_excel_mes, engine="openpyxl") as writer:
                                df_exportar_mes.to_excel(
                                    writer,
                                    index=False,
                                    sheet_name=f"{nome_mes} {ano_mes}"[:31]
                                )

                            st.session_state[chave_excel_pronto] = buffer_excel_mes.getvalue()

                        if st.session_state.get(chave_excel_pronto):

                            st.download_button(
                                f"⬇️ Baixar {nome_mes}/{ano_mes}.xlsx",
                                data=st.session_state[chave_excel_pronto],
                                file_name=f"analise_tecnica_luxiz_{chave_mes}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key=f"exportar_mes_{chave_mes}"
                            )

        _fragmento_sac()

    # =====================================================
    # AUDITORIA DE ATIVIDADES
    # =====================================================

    with tab_auditoria:

        @st.fragment
        def _fragmento_auditoria():

            estilos.exibir_notificacao_pendente()


            st.subheader(
                "🎯 Registrar Auditoria de Atividades"
            )

            st.caption(
                "O registro aparece automaticamente no card da pessoa "
                "informada no campo Nome."
            )

            FUNCOES_AUDITORIA = [
                "Conferente",
                "Empilhador",
                "Assistente Logístico"
            ]

            with st.form(
                "form_auditoria",
                clear_on_submit=True
            ):

                col1, col2 = st.columns(2)

                with col1:
                    nome_auditoria = st.text_input(
                        "Nome"
                    )

                with col2:
                    funcao_auditoria = st.selectbox(
                        "Função",
                        FUNCOES_AUDITORIA
                    )

                col3, col4, col5 = st.columns(3)

                with col3:
                    qtd_acertos_auditoria = st.number_input(
                        "Quantidade de acertos",
                        min_value=0,
                        step=1
                    )

                with col4:
                    qtd_erros_auditoria = st.number_input(
                        "Quantidade de erros",
                        min_value=0,
                        step=1
                    )

                with col5:
                    data_auditoria = st.date_input(
                        "Data"
                    )

                descricao_auditoria = st.text_area(
                    "Descrição"
                )

                registrar_auditoria = st.form_submit_button(
                    "➕ Registrar Auditoria"
                )

                if registrar_auditoria:

                    if not nome_auditoria:

                        st.error(
                            "Informe o nome da pessoa auditada."
                        )

                    else:

                        with estilos.mostrar_processando(f"registrando auditoria de {nome_auditoria}..."):
                            banco.adicionar_auditoria(
                                nome_auditoria,
                                funcao_auditoria,
                                qtd_acertos_auditoria,
                                qtd_erros_auditoria,
                                data_auditoria,
                                descricao_auditoria,
                                armazem_id_atual,
                                usuario=usuario_logado
                            )

                        estilos.notificar_sucesso(f"auditoria de {nome_auditoria} registrada.")
                        st.rerun(scope="fragment")

            st.divider()

            registros_auditoria = banco.ler_auditoria(armazem_id_atual)

            if not registros_auditoria:

                st.info(
                    "Nenhum registro de auditoria cadastrado ainda."
                )

            else:

                with st.expander(
                    f"📋 Ver registros de auditoria ({len(registros_auditoria)})",
                    expanded=False
                ):

                    st.caption(
                        "Marque as caixinhas para excluir vários de uma vez, "
                        "ou clique em ❌ para excluir um lançamento só:"
                    )

                    ids_selecionados_auditoria = []

                    for registro in registros_auditoria:

                        c0, c1, c2 = st.columns([0.6, 7.4, 1])

                        with c0:

                            marcado_auditoria = st.checkbox(
                                "selecionar",
                                key=f"select_auditoria_{registro['id']}",
                                label_visibility="collapsed"
                            )

                            if marcado_auditoria:
                                ids_selecionados_auditoria.append(registro["id"])

                        with c1:

                            st.caption(
                                f"👤 {registro['nome']} • {registro['funcao']} • "
                                f"✅ {registro['qtd_acertos']} / ❌ {registro['qtd_erros']} • "
                                f"{registro['data_atividade'].strftime('%d/%m/%Y')}"
                            )

                        with c2:

                            with st.popover("❌", key=f"pop_del_auditoria_{registro['id']}"):

                                confirmar_exclusao_auditoria(
                                    registro["id"],
                                    registro["nome"],
                                    armazem_id_atual
                                )

                    if ids_selecionados_auditoria:

                        st.write("")

                        with st.popover(
                            f"🗑️ Excluir {len(ids_selecionados_auditoria)} selecionado(s)",
                            key="pop_excluir_lote_auditoria"
                        ):

                            confirmar_exclusao_multipla_auditoria(
                                ids_selecionados_auditoria,
                                armazem_id_atual
                            )

                    st.divider()

                    df_exportar_auditoria = pd.DataFrame(registros_auditoria)

                    buffer_excel_auditoria = io.BytesIO()

                    with pd.ExcelWriter(buffer_excel_auditoria, engine="openpyxl") as writer:
                        df_exportar_auditoria.to_excel(
                            writer,
                            index=False,
                            sheet_name="Auditoria"
                        )

                    st.download_button(
                        "📥 Exportar para Excel",
                        data=buffer_excel_auditoria.getvalue(),
                        file_name="auditoria_atividades_luxiz.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="btn_exportar_auditoria"
                    )

        _fragmento_auditoria()

    # =====================================================
    # RODÍZIO - FIM DE EXPEDIENTE
    # =====================================================

    with tab_rotativo:

        @st.fragment
        def _fragmento_rotativo():

            estilos.exibir_notificacao_pendente()


            st.subheader(
                "🔄 Rodízio de Fim de Expediente"
            )

            st.caption(
                "Cadastre as pessoas e as atividades — o sistema monta "
                "a escala da semana sozinho, na aba 🔄 Rodízio."
            )

            col_pessoas, col_atividades = st.columns(2)

            # -----------------------------------------------
            # PESSOAS
            # -----------------------------------------------

            with col_pessoas:

                st.markdown(
                    "#### 👥 Pessoas no rodízio"
                )

                with st.form(
                    "form_pessoa_rotativo",
                    clear_on_submit=True
                ):

                    nome_pessoa_rotativo = st.text_input(
                        "Nome"
                    )

                    adicionar_pessoa = st.form_submit_button(
                        "➕ Adicionar pessoa"
                    )

                    if adicionar_pessoa:

                        if not nome_pessoa_rotativo:

                            st.error(
                                "Informe o nome da pessoa."
                            )

                        else:

                            with estilos.mostrar_processando(f"adicionando '{nome_pessoa_rotativo}'..."):
                                banco.adicionar_pessoa_rotativo(
                                    nome_pessoa_rotativo,
                                    armazem_id_atual
                                )

                            estilos.notificar_sucesso(f"'{nome_pessoa_rotativo}' adicionado(a) ao rodízio.")
                            st.rerun(scope="fragment")

                pessoas_rotativo = banco.listar_pessoas_rotativo(armazem_id_atual)

                if not pessoas_rotativo:

                    st.info(
                        "Nenhuma pessoa cadastrada ainda."
                    )

                else:

                    with st.expander(
                        f"👥 Ver pessoas cadastradas ({len(pessoas_rotativo)})",
                        expanded=False
                    ):

                        for id_pessoa, nome_pessoa in pessoas_rotativo:

                            c1, c2 = st.columns([8, 1])

                            with c1:
                                st.write(nome_pessoa)

                            with c2:

                                with st.popover("❌", key=f"pop_del_pessoa_rot_{id_pessoa}"):

                                    confirmar_exclusao_pessoa_rotativo(
                                        id_pessoa,
                                        nome_pessoa,
                                        armazem_id_atual
                                    )

            # -----------------------------------------------
            # ATIVIDADES
            # -----------------------------------------------

            with col_atividades:

                st.markdown(
                    "#### 🧹 Atividades do rodízio"
                )

                tipo_atividade_rotativo = st.selectbox(
                    "Tipo de escala",
                    [
                        "Rotativo (entra no rodízio)",
                        "Fixo (pessoa não rotaciona)"
                    ],
                    key="tipo_atividade_rotativo_select"
                )

                eh_atividade_fixa = tipo_atividade_rotativo.startswith("Fixo")

                with st.form(
                    "form_atividade_rotativo",
                    clear_on_submit=True
                ):

                    nome_atividade_rotativo = st.text_input(
                        "Nome da atividade"
                    )

                    pessoa_fixa_atividade = None

                    if eh_atividade_fixa:

                        pessoa_fixa_atividade = st.text_input(
                            "Pessoa responsável (fixa)"
                        )

                    adicionar_atividade = st.form_submit_button(
                        "➕ Adicionar atividade"
                    )

                    if adicionar_atividade:

                        if not nome_atividade_rotativo:

                            st.error(
                                "Informe o nome da atividade."
                            )

                        elif eh_atividade_fixa and not pessoa_fixa_atividade:

                            st.error(
                                "Informe a pessoa responsável fixa."
                            )

                        else:

                            with estilos.mostrar_processando(f"adicionando '{nome_atividade_rotativo}'..."):
                                banco.adicionar_atividade_rotativo(
                                    nome_atividade_rotativo,
                                    armazem_id_atual,
                                    tipo="fixo" if eh_atividade_fixa else "rotativo",
                                    pessoa_fixa=pessoa_fixa_atividade if eh_atividade_fixa else None
                                )

                            estilos.notificar_sucesso(f"'{nome_atividade_rotativo}' adicionada ao rodízio.")
                            st.rerun(scope="fragment")

                atividades_rotativo = banco.listar_atividades_rotativo(armazem_id_atual)

                if not atividades_rotativo:

                    st.info(
                        "Nenhuma atividade cadastrada ainda."
                    )

                else:

                    with st.expander(
                        f"🧹 Ver atividades cadastradas ({len(atividades_rotativo)})",
                        expanded=False
                    ):

                        for id_atividade, nome_atividade, tipo_atividade, pessoa_fixa_cadastrada in atividades_rotativo:

                            c1, c2 = st.columns([8, 1])

                            with c1:

                                if tipo_atividade == "fixo":

                                    st.write(
                                        f"📌 {nome_atividade} — fixo com "
                                        f"{pessoa_fixa_cadastrada or '?'}"
                                    )

                                else:

                                    st.write(
                                        f"🔄 {nome_atividade} — rotativo"
                                    )

                            with c2:

                                with st.popover("❌", key=f"pop_del_ativ_rot_{id_atividade}"):

                                    confirmar_exclusao_atividade_rotativo(
                                        id_atividade,
                                        nome_atividade,
                                        armazem_id_atual
                                    )

        _fragmento_rotativo()

    # =====================================================
    # EQUIPAMENTOS
    # =====================================================

    with tab_equipamentos:

        @st.fragment
        def _fragmento_equipamentos():

            estilos.exibir_notificacao_pendente()


            st.subheader(
                "🧰 Equipamentos"
            )

            st.caption(
                "Responsáveis por hidráulicos e carrinhos, e carrinhos fixos por local."
            )

            st.divider()

            # -----------------------------------------------
            # RESPONSÁVEIS POR HIDRÁULICOS
            # -----------------------------------------------

            st.markdown("#### 🔧 Responsáveis por Hidráulicos")

            with st.form("form_resp_hidraulico", clear_on_submit=True):

                col1, col2 = st.columns(2)

                with col1:
                    nome_resp_hid = st.text_input(
                        "Nome",
                        key="nome_resp_hid"
                    )

                with col2:
                    numero_resp_hid = st.text_input(
                        "Número do Hidráulico",
                        key="numero_resp_hid"
                    )

                if st.form_submit_button("➕ Adicionar Responsável"):

                    if not nome_resp_hid or not numero_resp_hid:

                        st.error(
                            "Preencha o Nome e o Número antes de adicionar."
                        )

                    else:

                        with estilos.mostrar_processando("adicionando responsável..."):
                            banco.adicionar_responsavel_hidraulico(
                                nome_resp_hid,
                                numero_resp_hid,
                                armazem_id_atual
                            )

                        estilos.notificar_sucesso("responsável adicionado.")
                        st.rerun(scope="fragment")

            responsaveis_hidraulicos = banco.ler_responsaveis_hidraulicos(armazem_id_atual)

            if responsaveis_hidraulicos:

                with st.expander(
                    f"🔧 Ver responsáveis por Hidráulicos ({len(responsaveis_hidraulicos)})",
                    expanded=False
                ):

                    for item in responsaveis_hidraulicos:

                        c1, c2, c3 = st.columns([7, 1, 1])

                        with c1:
                            st.write(
                                f"**{item['nome']}** — Hidráulico {item['numero']}"
                            )

                        with c2:

                            with st.popover("✏️", key=f"pop_edit_resp_hid_{item['id']}"):

                                st.markdown("**Editar responsável**")

                                novo_nome = st.text_input(
                                    "Nome",
                                    value=item["nome"],
                                    key=f"edit_nome_resp_hid_{item['id']}"
                                )

                                novo_numero = st.text_input(
                                    "Número do Hidráulico",
                                    value=item["numero"],
                                    key=f"edit_numero_resp_hid_{item['id']}"
                                )

                                if st.button(
                                    "💾 Salvar",
                                    width='stretch',
                                    key=f"salvar_edit_resp_hid_{item['id']}"
                                ):
                                    with estilos.mostrar_processando("salvando alterações..."):
                                        banco.editar_responsavel_hidraulico(
                                            item["id"],
                                            novo_nome,
                                            novo_numero,
                                            armazem_id_atual
                                        )
                                    estilos.notificar_sucesso("responsável atualizado.")
                                    st.rerun(scope="fragment")

                        with c3:

                            with st.popover("❌", key=f"pop_del_resp_hid_{item['id']}"):

                                confirmar_exclusao_responsavel_hidraulico(
                                    item["id"],
                                    item["nome"],
                                    item["numero"],
                                    armazem_id_atual
                                )

            else:

                st.info("Nenhum responsável cadastrado ainda.")

            st.divider()

            # -----------------------------------------------
            # RESPONSÁVEIS POR CARRINHOS
            # -----------------------------------------------

            st.markdown("#### 🛒 Responsáveis por Carrinhos")

            with st.form("form_resp_carrinho", clear_on_submit=True):

                col1, col2 = st.columns(2)

                with col1:
                    nome_resp_car = st.text_input(
                        "Nome",
                        key="nome_resp_car"
                    )

                with col2:
                    numero_resp_car = st.text_input(
                        "Número do Carrinho",
                        key="numero_resp_car"
                    )

                if st.form_submit_button("➕ Adicionar Responsável"):

                    if not nome_resp_car or not numero_resp_car:

                        st.error(
                            "Preencha o Nome e o Número antes de adicionar."
                        )

                    else:

                        with estilos.mostrar_processando("adicionando responsável..."):
                            banco.adicionar_responsavel_carrinho(
                                nome_resp_car,
                                numero_resp_car,
                                armazem_id_atual
                            )

                        estilos.notificar_sucesso("responsável adicionado.")
                        st.rerun(scope="fragment")

            responsaveis_carrinhos = banco.ler_responsaveis_carrinhos(armazem_id_atual)

            if responsaveis_carrinhos:

                with st.expander(
                    f"🛒 Ver responsáveis por Carrinhos ({len(responsaveis_carrinhos)})",
                    expanded=False
                ):

                    for item in responsaveis_carrinhos:

                        c1, c2, c3 = st.columns([7, 1, 1])

                        with c1:
                            st.write(
                                f"**{item['nome']}** — Carrinho {item['numero']}"
                            )

                        with c2:

                            with st.popover("✏️", key=f"pop_edit_resp_car_{item['id']}"):

                                st.markdown("**Editar responsável**")

                                novo_nome = st.text_input(
                                    "Nome",
                                    value=item["nome"],
                                    key=f"edit_nome_resp_car_{item['id']}"
                                )

                                novo_numero = st.text_input(
                                    "Número do Carrinho",
                                    value=item["numero"],
                                    key=f"edit_numero_resp_car_{item['id']}"
                                )

                                if st.button(
                                    "💾 Salvar",
                                    width='stretch',
                                    key=f"salvar_edit_resp_car_{item['id']}"
                                ):
                                    with estilos.mostrar_processando("salvando alterações..."):
                                        banco.editar_responsavel_carrinho(
                                            item["id"],
                                            novo_nome,
                                            novo_numero,
                                            armazem_id_atual
                                        )
                                    estilos.notificar_sucesso("responsável atualizado.")
                                    st.rerun(scope="fragment")

                        with c3:

                            with st.popover("❌", key=f"pop_del_resp_car_{item['id']}"):

                                confirmar_exclusao_responsavel_carrinho(
                                    item["id"],
                                    item["nome"],
                                    item["numero"],
                                    armazem_id_atual
                                )

            else:

                st.info("Nenhum responsável cadastrado ainda.")

            st.divider()

            # -----------------------------------------------
            # CARRINHOS FIXOS POR LOCAL
            # -----------------------------------------------

            st.markdown("#### 📍 Carrinhos Fixos por Local")

            st.caption(
                "Carrinhos que já ficam disponíveis fixos em cada local "
                "(ex: Remanejamento, Fracionado), sem precisar de remanejamento."
            )

            carrinhos_fixos = banco.ler_carrinhos_fixos(armazem_id_atual)

            locais_existentes = sorted({
                item["local"] for item in carrinhos_fixos
            }) or ["Remanejamento", "Fracionado"]

            with st.form("form_carrinho_fixo", clear_on_submit=True):

                col1, col2 = st.columns(2)

                with col1:
                    local_selecionado = st.selectbox(
                        "Local",
                        locais_existentes,
                        key="local_carrinho_fixo"
                    )

                    novo_local = st.text_input(
                        "Ou digite um novo local (opcional)",
                        key="novo_local_carrinho_fixo"
                    )

                with col2:
                    numero_carrinho_fixo = st.text_input(
                        "Número do Carrinho",
                        key="numero_carrinho_fixo"
                    )

                local_carrinho_fixo = novo_local.strip() or local_selecionado

                if st.form_submit_button("➕ Adicionar Carrinho Fixo"):

                    if not local_carrinho_fixo or not numero_carrinho_fixo:

                        st.error(
                            "Preencha o Local e o Número antes de adicionar."
                        )

                    else:

                        with estilos.mostrar_processando("adicionando carrinho fixo..."):
                            banco.adicionar_carrinho_fixo(
                                local_carrinho_fixo,
                                numero_carrinho_fixo,
                                armazem_id_atual
                            )

                        estilos.notificar_sucesso("carrinho fixo adicionado.")
                        st.rerun(scope="fragment")

            if carrinhos_fixos:

                with st.expander(
                    f"📍 Ver carrinhos fixos por local ({len(carrinhos_fixos)})",
                    expanded=False
                ):

                    por_local = {}

                    for item in carrinhos_fixos:
                        por_local.setdefault(item["local"], []).append(item)

                    for local in sorted(por_local.keys()):

                        st.write(f"**📍 {local}**")

                        for item in por_local[local]:

                            c1, c2, c3 = st.columns([7, 1, 1])

                            with c1:
                                st.caption(f"Carrinho {item['numero']}")

                            with c2:

                                with st.popover("✏️", key=f"pop_edit_carfixo_{item['id']}"):

                                    st.markdown("**Editar carrinho fixo**")

                                    novo_local = st.text_input(
                                        "Local",
                                        value=item["local"],
                                        key=f"edit_local_carfixo_{item['id']}"
                                    )

                                    novo_numero = st.text_input(
                                        "Número do Carrinho",
                                        value=item["numero"],
                                        key=f"edit_numero_carfixo_{item['id']}"
                                    )

                                    if st.button(
                                        "💾 Salvar",
                                        width='stretch',
                                        key=f"salvar_edit_carfixo_{item['id']}"
                                    ):
                                        with estilos.mostrar_processando("salvando alterações..."):
                                            banco.editar_carrinho_fixo(
                                                item["id"],
                                                novo_local,
                                                novo_numero,
                                                armazem_id_atual
                                            )
                                        estilos.notificar_sucesso("carrinho fixo atualizado.")
                                        st.rerun(scope="fragment")

                            with c3:

                                with st.popover("❌", key=f"pop_del_carfixo_{item['id']}"):

                                    confirmar_exclusao_carrinho_fixo(
                                        item["id"],
                                        item["local"],
                                        item["numero"],
                                        armazem_id_atual
                                    )

            else:

                st.info("Nenhum carrinho fixo cadastrado ainda.")

        _fragmento_equipamentos()

    # =====================================================
    # USUÁRIOS
    # =====================================================

    BADGES_USUARIO = {
        "Fundador.": ("👑", "Fundador", "#f59e0b"),
        "Gestao.": ("🛡️", "Gestão", "#3b82f6"),
        "Painel.": ("📟", "Painel Logístico", "#6366f1"),
        "Separador.": ("📦", "Separador", "#22c55e"),
        "Conferente.": ("🔎", "Conferente", "#06b6d4"),
        "Recebimento.": ("📥", "Recebimento", "#a855f7"),
        "Empilhador.": ("🏗️", "Empilhador", "#ec4899"),
        "Assistente.": ("🧑‍💼", "Assistente Logístico", "#64748b"),
    }

    def badge_usuario(nome):

        nome = nome or ""

        for prefixo, dados in BADGES_USUARIO.items():

            if nome.startswith(prefixo):
                return dados

        return ("👤", "Usuário", "#64748b")

    def status_presenca(ultimo_acesso):

        if not ultimo_acesso:
            return "⚪", "Nunca acessou ainda", None

        agora_utc = datetime.now(timezone.utc)
        acesso_utc = ultimo_acesso.replace(tzinfo=timezone.utc)

        segundos = (agora_utc - acesso_utc).total_seconds()

        horario_local = acesso_utc.astimezone(
            ZoneInfo("America/Campo_Grande")
        )

        horario_formatado = horario_local.strftime("%d/%m/%Y %H:%M")

        if segundos <= 240:
            return "🟢", "Online agora", horario_formatado

        minutos = int(segundos // 60)
        horas = minutos // 60
        dias = horas // 24

        if dias >= 1:
            tempo_texto = f"há {dias} dia(s)"
        elif horas >= 1:
            tempo_texto = f"há {horas}h"
        elif minutos >= 1:
            tempo_texto = f"há {minutos} min"
        else:
            tempo_texto = "há poucos segundos"

        return "⚪", f"Visto {tempo_texto}", horario_formatado

    if admin_master:

        with tab_usuarios:

            @st.fragment
            def _fragmento_usuarios():

                estilos.exibir_notificacao_pendente()


                estilos.cabecalho_pagina(
                    "👥",
                    "Gerenciamento de Usuários",
                    "Cadastro de acessos e visão de quem está no sistema agora.",
                    cor="#8b5cf6"
                )

                usuarios = banco.listar_usuarios(armazem_id_atual)

                # =====================================================
                # SEPARA REGISTROS REAIS DE REGISTROS INCONSISTENTES
                # =====================================================
                # "Inconsistente" = linha sem nome de usuário (sobra de
                # algum cadastro que falhou no passado). Cada um desses
                # antes virava um card individual — com muitos deles
                # cadastrados, isso pesava na tela à toa. Agora só os
                # usuários de verdade viram card; os inconsistentes
                # entram num resumo único, com exclusão em lote.

                usuarios_validos = [u for u in usuarios if u[1]]
                usuarios_inconsistentes = [u for u in usuarios if not u[1]]

                total_usuarios = len(usuarios_validos)

                total_online = sum(
                    1 for u in usuarios_validos
                    if u[3] and (
                        datetime.now(timezone.utc)
                        - u[3].replace(tzinfo=timezone.utc)
                    ).total_seconds() <= 240
                )

                c1, c2 = st.columns(2)

                with c1:
                    st.metric("👥 Usuários cadastrados", total_usuarios)

                with c2:
                    st.metric("🟢 Online agora", total_online)

                st.divider()

                if usuarios_inconsistentes:

                    with st.container(border=True):

                        st.warning(
                            f"⚠️ {len(usuarios_inconsistentes)} registro(s) "
                            "inconsistente(s) encontrado(s) (usuários sem "
                            "nome, provavelmente de cadastros que falharam). "
                            "Não contam como usuário e não geram card — "
                            "recomendo excluir."
                        )

                        with st.popover("🗑️ Excluir todos os inconsistentes"):

                            confirmar_exclusao_usuarios_inconsistentes(
                                [u[0] for u in usuarios_inconsistentes]
                            )

                    st.divider()

                with st.expander("➕ Criar novo usuário"):

                    if fundador:

                        st.caption(
                            f"📍 Os usuários criados abaixo serão vinculados a: "
                            f"**{st.session_state.get('armazem_visualizado_nome', '—')}**"
                        )

                    novo_usuario = st.text_input(
                        "Usuário"
                    )

                    senha_usuario = st.text_input(
                        "Senha Inicial",
                        type="password"
                    )

                    if st.button(
                        "➕ Criar Usuário"
                    ):

                        if not novo_usuario.strip() or not senha_usuario.strip():

                            st.warning(
                                "Preencha o usuário e a senha inicial."
                            )

                        else:

                            try:

                                with estilos.mostrar_processando(f"criando usuário '{novo_usuario.strip()}'..."):
                                    banco.criar_usuario(
                                        novo_usuario.strip(),
                                        senha_usuario,
                                        armazem_id_atual
                                    )

                                estilos.notificar_sucesso(f"usuário '{novo_usuario.strip()}' criado.")
                                st.rerun(scope="fragment")

                            except Exception as erro:

                                st.error(
                                    str(erro)
                                )

                st.divider()

                st.markdown("##### 👥 Usuários cadastrados")

                if not usuarios_validos:

                    st.info("Nenhum usuário cadastrado ainda.")

                else:

                    cols = st.columns(3)

                    for indice, usuario in enumerate(usuarios_validos):

                        uid = usuario[0]
                        nome = usuario[1]
                        ultimo_acesso = usuario[3]

                        emblema, rotulo_funcao, cor = badge_usuario(nome)
                        emblema_status, texto_status, horario_acesso = status_presenca(ultimo_acesso)

                        nome_exibicao = (
                            nome.split(".", 1)[1].strip().title()
                            if "." in nome
                            else nome
                        )

                        chave_card = f"card-user-{uid}"

                        st.markdown(
                            f"""
                            <style>
                            .st-key-{chave_card} {{
                                background-color: {cor}22 !important;
                                border: 2px solid {cor} !important;
                                border-radius: 0.8rem;
                            }}
                            </style>
                            """,
                            unsafe_allow_html=True
                        )

                        with cols[indice % 3]:

                            with st.container(border=True, key=chave_card):

                                st.markdown(
                                    f"""
                                    <div style="
                                        width:52px;height:52px;border-radius:50%;
                                        background:{cor};
                                        display:flex;align-items:center;justify-content:center;
                                        font-size:1.5rem;margin-bottom:.4rem;
                                        box-shadow:0 0 14px {cor}88;
                                    ">{emblema}</div>
                                    """,
                                    unsafe_allow_html=True
                                )

                                st.markdown(f"### {nome_exibicao}")

                                st.caption(rotulo_funcao)

                                st.markdown(
                                    f"{emblema_status} {texto_status}"
                                )

                                if horario_acesso:

                                    st.caption(
                                        f"Último acesso: {horario_acesso}"
                                    )

                                if nome == "Fundador.henrique":

                                    st.write("")
                                    st.caption("👑 Conta principal — não pode ser excluída.")

                                    with st.popover("🪪 Perfil", key=f"pop_perfil_user_{uid}"):

                                        perfil._formulario_perfil(
                                            nome,
                                            armazem_id_atual,
                                            chave_prefixo=f"admcard_{uid}_"
                                        )

                                else:

                                    col_perfil, col_excluir = st.columns(2)

                                    with col_perfil:

                                        with st.popover("🪪 Perfil", key=f"pop_perfil_user_{uid}"):

                                            perfil._formulario_perfil(
                                                nome,
                                                armazem_id_atual,
                                                chave_prefixo=f"admcard_{uid}_"
                                            )

                                    with col_excluir:

                                        with st.popover("🗑️ Excluir", key=f"pop_del_user_{uid}"):

                                            confirmar_exclusao_usuario(
                                                uid,
                                                nome or nome_exibicao
                                            )

            _fragmento_usuarios()

    # =====================================================
    # ARMAZÉNS (SÓ FUNDADOR)
    # =====================================================

    if fundador:

        with tab_armazens:

            @st.fragment
            def _fragmento_armazens():

                estilos.exibir_notificacao_pendente()


                st.subheader(
                    "🏢 Armazéns Cadastrados"
                )

                st.caption(
                    "Cada armazém tem seus próprios dados, totalmente "
                    "separados dos demais. Use o seletor '📍 Visualizando "
                    "dados de' no topo do app para trocar entre eles."
                )

                lista_armazens = banco.listar_armazens()

                for id_armazem, nome_armazem in lista_armazens:

                    c1, c2 = st.columns([4, 1])

                    with c1:

                        novo_nome = st.text_input(
                            "Nome",
                            value=nome_armazem,
                            key=f"nome_armazem_{id_armazem}",
                            label_visibility="collapsed"
                        )

                    with c2:

                        if st.button(
                            "💾 Salvar",
                            key=f"salvar_armazem_{id_armazem}"
                        ):

                            if novo_nome.strip() and novo_nome != nome_armazem:

                                banco.renomear_armazem(
                                    id_armazem,
                                    novo_nome.strip()
                                )

                                estilos.notificar_sucesso(
                                    f"✨ Luxiz IA: armazém renomeado para "
                                    f"'{novo_nome.strip()}'."
                                )
                                st.rerun(scope="fragment")

                st.divider()

                st.subheader(
                    "➕ Cadastrar Novo Armazém"
                )

                st.caption(
                    "Cria um novo cliente/armazém com dados totalmente "
                    "isolados. Depois, crie o primeiro usuário dele na "
                    "aba '👥 Usuários' (selecione esse armazém no seletor "
                    "do topo antes de criar o usuário)."
                )

                nome_novo_armazem = st.text_input(
                    "Nome do novo armazém"
                )

                if st.button(
                    "➕ Criar Armazém"
                ):

                    if not nome_novo_armazem.strip():

                        st.error(
                            "Informe um nome para o novo armazém."
                        )

                    else:

                        try:

                            novo_id_armazem = banco.criar_armazem(
                                nome_novo_armazem.strip()
                            )

                            estilos.notificar_sucesso(
                                f"✨ Luxiz IA: armazém "
                                f"'{nome_novo_armazem.strip()}' criado."
                            )

                            st.session_state.armazem_visualizado_id = novo_id_armazem
                            st.session_state.armazem_visualizado_nome = nome_novo_armazem.strip()

                            st.rerun(scope="fragment")

                        except Exception as erro:

                            st.error(
                                str(erro)
                            )
            _fragmento_armazens()